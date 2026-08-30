#!/usr/bin/env python3
"""Export the FULL results deck — extraction + sorter sweep + LegalBench —
as a Google-Slides-style xlsx.

One workbook, every worksheet a 16:9 "slide" (landscape, fit-to-page, dark
title banner + footer) so the sheets read as a deck and copy-paste cleanly
into Google Slides for continued editing:

- **Part A — Extraction (contracts specialist)** — the FULL run lineage
  (all 61 collected extraction/specialist records, chronological, with
  overall / field-presence / schema-valid / verified-precision / tokens /
  cost), the champion v32 ``510_full_clean`` detail (metadata, headlines +
  v31 comparison, per-field scores, error decomposition, entity-list F1,
  MAE/R² diagnostics), and the full extraction codebook (fields + types,
  scoring rubric & thresholds).
- **Part B — Sorter subtype sweep** — ALL collected ``sorter_v13``
  subtype-classification runs (8 models + smokes/reruns, chronological,
  strict/equiv/exact/fails), the champion detail with failure-mode
  breakdown + top-confidence failure examples, and the full sorter
  codebook (25 CUAD subtypes, equivalence families, scoring rules).
- **Part C — LegalBench** — the complete performance log (all 56 collected
  LegalBench records across task_v0/v1/v2 hearsay + task_v3/v4 contract
  families, with n / exact / CI / per-class accuracy / tokens / cost),
  a per-task results summary, and a task legend.

Data is read live from ``reports/experiment_log.jsonl``,
``config/taxonomy.yaml`` and ``agents/sorter_agent.py`` — no network, no LLM.

Usage::

    python scripts/reporting/export_full_results_deck.py
    python scripts/reporting/export_full_results_deck.py --outdir reports/sheets
    python scripts/reporting/export_full_results_deck.py --log reports/experiment_log.jsonl
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import datetime, timezone
from typing import Any, Callable

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# Reuse the slide geometry + table styling from the existing deck builder.
from export_slides_deck import (  # noqa: E402
    MODEL_DISPLAY,
    _BOX,
    _GRAY,
    _HILITE,
    _NAVY,
    _ROW_ALT,
    _model,
    _num,
    _pct,
    _trunc,
    big_number,
    kv_table,
    load_field_scoring,
    load_records_list,
    note,
    section_row,
    table,
)

try:
    import yaml  # type: ignore

    _HAS_YAML = True
except Exception:  # pragma: no cover - yaml is an optional nicety
    _HAS_YAML = False

from agents.sorter_agent import (  # noqa: E402
    CONTRACT_SUBTYPES,
    SUBTYPE_EQUIVALENCES,
)

DEFAULT_LOG = "reports/experiment_log.jsonl"
DEFAULT_TAXONOMY = "config/taxonomy.yaml"
DEFAULT_OUTDIR = "reports/sheets"
DEFAULT_DECK = "extraction_sweep_legalbench_full_deck.xlsx"

_TOTAL_SLIDES = 19

# Champion runs pinned for the detail slides.
EXTRACTION_CHAMPION = "qwen3.7-flash_contracts_specialist_v32_extraction_langfuse_510_full_clean"
SORTER_CHAMPION = "qwen3.7-flash_sorter_v13_subtype_langfuse@2026-08-16T03:56:38"

_WHITE = "FFFFFF"
_BLUE = "2E75B6"
_GREEN = "E2EFDA"
_RED = "FCE4EC"
_THIN = Side(style="thin", color="BFBFBF")
_BOX_ALL = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_N_COLS = 16
_N_ROWS = 36


# ---------------------------------------------------------------------------
# Slide geometry / styling (one worksheet = one 16:9 slide)
# ---------------------------------------------------------------------------


def _widths(ws: openpyxl.worksheet.worksheet.Worksheet, widths: list[float]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def new_slide(
    wb: openpyxl.Workbook,
    tag: str,
    title: str,
    n: int,
    kicker: str = "",
    widths: list[float] | None = None,
) -> tuple[openpyxl.worksheet.worksheet.Worksheet, int]:
    """Create a slide sheet with the banner + footer; returns (ws, cursor_row)."""
    short_tags = {
        "PART A · EXTRACTION": "EXTRACTION",
        "PART B · SORTER SWEEP": "SORTER SWEEP",
        "PART C · LEGALBENCH": "LEGALBENCH",
    }
    name = f"{n:02d}. {short_tags.get(tag, tag)} · {title}"
    if n == 1:
        name = "01. COVER"
    ws = wb.create_sheet(name[:31])
    _widths(ws, widths or [2, 21, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 2])
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    from openpyxl.worksheet.properties import PageSetupProperties

    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins.left = ws.page_margins.right = 0.3
    ws.page_margins.top = ws.page_margins.bottom = 0.3

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=_N_COLS)
    banner = ws.cell(row=1, column=1, value=f"{tag}  ·  {title}")
    banner.font = Font(bold=True, size=15, color=_WHITE)
    banner.fill = PatternFill("solid", fgColor=_NAVY)
    banner.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 46.0

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=_N_COLS)
    sub = ws.cell(row=2, column=1, value=kicker)
    sub.font = Font(size=9, italic=True, color=_GRAY)
    sub.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 16.0

    ws.merge_cells(start_row=_N_ROWS, start_column=1, end_row=_N_ROWS, end_column=_N_COLS)
    foot = ws.cell(
        row=_N_ROWS,
        column=1,
        value=f"{DEFAULT_DECK} · slide {n}/{_TOTAL_SLIDES} · "
        f"generated {datetime.now(timezone.utc).strftime('%Y-%m-%d')} · "
        f"source: reports/experiment_log.jsonl",
    )
    foot.font = Font(size=8, color=_GRAY)
    foot.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    foot.border = Border(top=Side(style="medium", color=_GRAY))
    ws.row_dimensions[_N_ROWS].height = 16.0
    return ws, 4


# ---------------------------------------------------------------------------
# Record helpers
# ---------------------------------------------------------------------------


def _get(record: dict, path: str, default: Any = None) -> Any:
    cur: Any = record
    for part in path.split("."):
        if isinstance(cur, dict) and part in cur:
            cur = cur[part]
        else:
            return default
    return cur


def _date(r: dict) -> str:
    return str(r.get("timestamp", ""))[:10]


def _short_run(name: str, n: int = 38) -> str:
    for prefix in ("qwen3.7-flash_", "deepseek-v4-flash_", "deepseek-v4-pro_",
                   "meta-llama-llama-3.3-70b-instruct_", "llama-4-scout_",
                   "gpt-5-nano_", "gpt-4.1-nano_", "gpt-4o-mini_"):
        if name.startswith(prefix):
            name = name[len(prefix):]
            break
    return _trunc(name, n)


def _mtok(r: dict) -> str:
    tot = (r.get("tokens") or {}).get("total_tokens") or 0
    return f"{tot / 1e6:.1f}M" if tot else "—"


def _cost(r: dict) -> str:
    c = (r.get("tokens") or {}).get("cost_estimated_usd")
    return f"${c:.3f}" if c is not None else "—"


def _short_model(r: dict) -> str:
    return _model(r)


def load_contract_fields(taxonomy_path: str) -> dict[str, str]:
    if not _HAS_YAML:
        return {}
    with open(taxonomy_path, encoding="utf-8") as fh:
        tax = yaml.safe_load(fh)
    for cls in tax.get("doc_classes", []):
        if cls.get("key") == "contract":
            return dict(cls.get("field_types", {}))
    return {}


# ---------------------------------------------------------------------------
# COVER + CONTENTS
# ---------------------------------------------------------------------------


def slide_cover(wb: openpyxl.Workbook) -> None:
    ws, _ = new_slide(wb, "COVER", "Full results deck — extraction + sorter sweep + LegalBench", 1)
    row = 8
    c = ws.cell(row=row, column=3, value="Contracts extraction  ·  Sorter subtype sweep  ·  LegalBench")
    c.font = Font(bold=True, size=22, color=_NAVY)
    row += 2
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=13)
    c = ws.cell(row=row, column=3, value="llm-entity-extraction — every collected run on the three tasks, as a slide deck")
    c.font = Font(size=12, italic=True, color=_GRAY)
    row += 3
    sub = [
        "Part A — Extraction · contracts_specialist v2→v32 lineage (61 runs) + champion v32 detail + extraction codebook",
        "Part B — Sorter sweep · all 22 sorter_v13 subtype runs (8 models) + champion detail + sorter codebook",
        "Part C — LegalBench · all 56 performance records (hearsay v0/v1/v2, contract families v3/v4) + summary + legend",
        "Every sheet is a 16:9 slide (landscape, fit-to-page) — import into Google Slides by copying each sheet.",
    ]
    for s in sub:
        c = ws.cell(row=row, column=3, value="•  " + s)
        c.font = Font(size=11)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[row].height = 22.0
        row += 1
    row += 2
    c = ws.cell(row=row, column=3, value="Sources: reports/experiment_log.jsonl · config/taxonomy.yaml · agents/sorter_agent.py · docs/SCORING.md")
    c.font = Font(size=9, italic=True, color=_GRAY)


def slide_contents(wb: openpyxl.Workbook) -> None:
    ws, row = new_slide(wb, "CONTENTS", "Deck map", 2)
    row = section_row(ws, row, "Part A — Extraction (contracts specialist, all collected runs)")
    row = note(ws, row, 2, "03–04 · Full run lineage (61 runs) · 05 · Champion v32 metadata & headlines (+ v31) · 06 · Per-field scores · 07 · Error decomposition & entity lists · 08 · Diagnostics MAE/R² · 09–10 · Extraction codebook (fields + rubric)")
    row = section_row(ws, row, "Part B — Sorter subtype sweep (sorter_v13, 8 models)")
    row = note(ws, row, 2, "11 · Full sweep — all 22 sorter_v13 subtype runs · 12 · Champion v13 headlines & failure modes · 13–14 · Sorter codebook (25 subtypes + rules)")
    row = section_row(ws, row, "Part C — LegalBench (all performance records)")
    row = note(ws, row, 2, "15 · Performance log — hearsay v0/v1/v2 · 16 · Performance log — task_v3 contract families · 17 · Performance log — task_v4 (+ sampled) · 18 · Per-task results summary · 19 · Task legend")
    row = note(ws, row + 1, 2, "Same-surface comparisons only: only runs on identical rows/surfaces are comparable (noise floor on the 50-doc chunked extraction surface ≈ ±0.03, on full-509 reruns ≈ 0.000).")

    return None


# ---------------------------------------------------------------------------
# PART A — EXTRACTION
# ---------------------------------------------------------------------------


def _extraction_records(records: list[dict]) -> list[dict]:
    return [r for r in records
            if "specialist" in r.get("experiment_name", "") or "extraction" in r.get("experiment_name", "")]


def _extraction_row(r: dict) -> list[Any]:
    s = r.get("scores", {})
    return [
        _date(r),
        _short_run(r.get("experiment_name", "")),
        _short_model(r),
        r.get("n_rows"),
        _num(s.get("overall_extraction_score"), 4) if s.get("overall_extraction_score") is not None else "—",
        _pct(s.get("field_presence")),
        _pct(s.get("schema_valid")),
        _pct(s.get("overall_verified_precision")),
        _mtok(r),
        _cost(r),
    ]


def slide_extract_lineage(wb: openpyxl.Workbook, records: list[dict], chunk: int, n: int) -> None:
    rows = _extraction_records(records)
    rows = sorted(rows, key=lambda r: r.get("timestamp", ""))
    half = 31
    part = rows[:half] if chunk == 0 else rows[half:]
    tag = "PART A · EXTRACTION"
    ws, row = new_slide(
        wb, tag, f"Run lineage — all collected runs ({len(rows)} total, part {chunk + 1}/2)",
        n,
        kicker="Every extraction / contracts-specialist record in reports/experiment_log.jsonl, chronological. "
               "— = not recorded for that run (chained pilots carry their own stage scores).",
        widths=[10, 40, 15, 8, 11, 11, 11, 13, 10, 10],
    )
    row = table(
        ws, row, 1,
        ["Date", "Run", "Model", "n", "Overall", "FieldPres", "SchemaV", "VerifPrec", "Tokens", "Cost est."],
        [_extraction_row(r) for r in part],
        fmt=None,
    )
    note(ws, row, 1, f"{len(part)} runs shown · champion = qwen3.7-flash × contracts_specialist_v32 (510-full-clean, 0.8807) · "
                     "full-510 surface runs are the only comparable ones (50-doc runs are a different surface).")


def slide_extract_champion_metadata(wb: openpyxl.Workbook, r: dict, n: int) -> None:
    ws, row = new_slide(
        wb, "PART A · EXTRACTION", "Champion v32 — metadata & headlines",
        n,
        kicker="qwen3.7-flash × contracts_specialist_v32, 510-full-clean — full-corpus chunked extraction (90k/8k windows).",
        widths=[2, 30, 22, 16, 16, 16, 16, 16, 16, 16, 16, 16, 16, 16, 16, 2],
    )
    meta = [
        ("Run", r.get("experiment_name")),
        ("Model", r.get("model")),
        ("Prompt", r.get("prompt_version")),
        ("Task", r.get("task")),
        ("Date", str(r.get("timestamp", ""))[:19]),
        ("Rows (n_ok / n_rows / n_error)", f"{r.get('n_ok')} / {r.get('n_rows')} / {r.get('n_error')}"),
        ("Tokens (prompt / completion / total)", _mtok_full(r)),
        ("Est. cost", _cost(r)),
        ("Chunking", "90k windows, 8k overlap (--chunked)"),
        ("Reasoning effort", str(_get(r, "parameters.reasoning_effort", "none"))),
        ("Tracing backend", str(_get(r, "parameters.tracing_backend", "—"))),
    ]
    row = kv_table(ws, row, 2, meta)
    row = section_row(ws, row, "Headline scores")
    s = r.get("scores", {})
    ci = s.get("overall_extraction_score_ci", {}) or {}
    row += 0
    big_number(ws, row, 2, _num(s.get("overall_extraction_score")), "overall extraction score")
    big_number(ws, row, 5, _pct(s.get("field_presence")), "field presence")
    big_number(ws, row, 8, _pct(s.get("schema_valid")), "schema valid")
    big_number(ws, row, 11, _pct(s.get("overall_verified_precision")), "verified precision")
    row += 3
    row = table(
        ws, row, 1,
        ["Run", "Overall", "FieldPres", "SchemaV", "VerifPrec", "Effective date"],
        [["contracts_specialist_v32 (510-full-clean)", _num(s.get("overall_extraction_score")),
          _pct(s.get("field_presence")), _pct(s.get("schema_valid")),
          _pct(s.get("overall_verified_precision")), _num(_get(s, "per_field.effective_date"))]],
        widths=[30, 12, 12, 12, 12, 16],
    )
    note(ws, row, 1, "v32 510-full-clean vs v31 510-full: Δ = +0.0070 overall (inside the ±0.011 chunked noise band — logic repair, not a claimed win; see docs/memos/contracts_specialist_v32…).")


def _mtok_full(r: dict) -> str:
    t = r.get("tokens") or {}
    return (f"{t.get('prompt_tokens', 0) / 1e6:.1f}M / {t.get('completion_tokens', 0) / 1e6:.1f}M / "
            f"{t.get('total_tokens', 0) / 1e6:.1f}M")


def slide_extract_per_field(wb: openpyxl.Workbook, r: dict, n: int) -> None:
    ws, row = new_slide(
        wb, "PART A · EXTRACTION", "Per-field scores",
        n,
        kicker="Champion v32 510-full-clean — deterministic field-type-aware scoring (see codebook slides).",
        widths=[2, 20, 12, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 2],
    )
    row = section_row(ws, row, "Per-field scores — score / verified precision / hallucination flag")
    pf = r.get("scores", {}).get("per_field", {}) or {}
    headers = ["Field", "Score", "Verified", "Halluc.", "Count", "Notes"]
    rows = []
    for fname, fvals in pf.items():
        if not isinstance(fvals, dict):
            rows.append([fname, _num(fvals), "—", "—", "—", "—"])
            continue
        rows.append([
            fname,
            _num(fvals.get("score")),
            _pct(fvals.get("verified_precision")),
            _num(fvals.get("hallucination_count") or 0),
            fvals.get("n") or _get(fvals, "n_scored") or "—",
            _trunc(fvals.get("notes") or fvals.get("failure_mode") or "", 60),
        ])
    rows.sort(key=lambda x: (x[1] == "—", str(x[1])))
    table(ws, row, 1, headers, rows, widths=[22, 12, 12, 12, 10, 40])
    note(ws, row, 1, "Entity-list fields (parties, key_obligations, termination_clauses) are scored by ground-truth coverage — "
                     "raw P/R/F1 stays in the entity-list breakdown slide. Containment fields (governing_law, term_length, renewal_terms) by expected-within-predicted token containment.")


def slide_extract_errors(wb: openpyxl.Workbook, r: dict, n: int) -> None:
    ws, row = new_slide(
        wb, "PART A · EXTRACTION", "Error decomposition & entity lists",
        n,
        kicker="Champion v32 510-full-clean — exact/partial/miss decomposition + list-field coverage.",
        widths=[2, 26, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 2],
    )
    d = r.get("scores", {})
    decomp = d.get("error_decomposition") or {}
    if decomp:
        row = section_row(ws, row, "Field-level error decomposition")
        table(ws, row, 1,
              ["Field", "Exact", "Partial", "Miss"],
              [[k, _num(v.get("exact")), _num(v.get("partial")), _num(v.get("miss"))]
               for k, v in decomp.items()],
              widths=[22, 12, 12, 12])
    row = section_row(ws, row, "Entity-list coverage-F1 vs raw P/R/F1 (partial-GT caveat)")
    els = d.get("entity_list_scores") or {}
    headers = ["Field", "GT coverage", "Precision", "Recall", "F1"]
    rows = [[k, _num(v.get("gt_coverage") or v.get("coverage")), _num(v.get("precision")),
             _num(v.get("recall")), _num(v.get("f1"))] for k, v in els.items()]
    table(ws, row, 1, headers, rows, widths=[22, 12, 12, 12, 12])
    note(ws, row, 1, "Lists are scored by ground-truth coverage (recall over matched labels), NOT F1 — CUAD clause-QA labels are "
                     "partial samples; correct extractions beyond the GT must not be penalized.")


def slide_extract_diagnostics(wb: openpyxl.Workbook, r: dict, n: int) -> None:
    ws, row = new_slide(
        wb, "PART A · EXTRACTION", "Regression diagnostics — MAE / R² / span drift",
        n,
        kicker="Champion v32 510-full-clean — run-level diagnostics from src/metrics.py (read MAE/R² with their support sizes).",
        widths=[2, 26, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 2],
    )
    diag = r.get("scores", {}).get("diagnostics") or {}
    row = section_row(ws, row, "MAE / R² (parseable-pair counts)")
    rows = []
    for kind in ("date", "duration", "money"):
        k = diag.get(f"{kind}_mae_days" if kind != "money" else "money_mae_usd") or {}
        if isinstance(k, dict):
            rows.append([kind, _num(k.get("mae")), _num(k.get("r2")), k.get("n_pairs"),
                         _num(k.get("median")), _num(k.get("pair_mae"))])
    if rows:
        table(ws, row, 1, ["Kind", "MAE", "R²", "n pairs", "Median", "Pair MAE"], rows,
              widths=[16, 12, 12, 12, 12, 12])
    row = section_row(ws, row, "Span-count drift (list fields)")
    drift = diag.get("span_count_drift") or {}
    if isinstance(drift, dict) and drift:
        table(ws, row, 1,
              ["Metric", "Value"],
              [[k, str(v)] for k, v in drift.items()],
              widths=[30, 30])
    note(ws, row, 1, "MAE/R² rows are only as good as their parseable-pair counts (date_n_pairs etc.) — always read the support size with the number.")


def slide_codebook_fields(wb: openpyxl.Workbook, fields: dict[str, str], n: int) -> None:
    ws, row = new_slide(
        wb, "PART A · EXTRACTION", "Codebook — extraction fields & types",
        n,
        kicker="config/taxonomy.yaml → contract doc-class field_types.",
        widths=[2, 24, 16, 12, 46, 16, 16, 16, 16, 16, 16, 16, 16, 16, 16, 2],
    )
    row = section_row(ws, row, "Fields, types & scoring class")
    table(ws, row, 1,
          ["Field", "Type", "Scoring class", "Notes"],
          [[k, (v.get("type") if isinstance(v, dict) else v), (v.get("scoring") if isinstance(v, dict) else "—"),
            _trunc(v.get("notes") if isinstance(v, dict) else "", 100)] for k, v in fields.items()],
          widths=[22, 14, 14, 60])


def slide_codebook_scoring(wb: openpyxl.Workbook, fs: dict[str, Any], n: int) -> None:
    ws, row = new_slide(
        wb, "PART A · EXTRACTION", "Codebook — scoring rubric & thresholds",
        n,
        kicker="config/taxonomy.yaml → field_scoring.",
        widths=[2, 40, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 2],
    )
    row = section_row(ws, row, "Scoring rules (field-type-aware)")
    flat = []
    if isinstance(fs, dict):
        for k, v in fs.items():
            if isinstance(v, dict):
                for k2, v2 in v.items():
                    flat.append([f"{k}.{k2}", _trunc(str(v2), 110)])
            else:
                flat.append([k, _trunc(str(v), 110)])
    table(ws, row, 1, ["Key", "Value"], flat, widths=[34, 80])
    note(ws, row, 1, "Ambiguous band [0.5, 0.85] triggers the optional --judge LLM pass. Entity lists via optimal bipartite "
                     "matching (Hungarian, threshold 0.6). Factuality guard: every predicted list item must match GT or be "
                     "grounded in the source (token coverage ≥ 0.7).")


# ---------------------------------------------------------------------------
# PART B — SORTER SWEEP
# ---------------------------------------------------------------------------


def _sweep_records(records: list[dict]) -> list[dict]:
    return [r for r in records if "_sorter_v13_subtype" in r.get("experiment_name", "")]


def _sweep_flag(r: dict, seen: dict[str, int]) -> str:
    name = r.get("experiment_name", "")
    key = name.split("@")[0]
    seen[key] = seen.get(key, 0) + 1
    n = seen[key]
    if name.endswith("smoke1") or name.endswith("smoke2"):
        return "smoke"
    if n > 1:
        return f"rerun #{n}"
    if r.get("scores", {}).get("sorter", {}).get("subtype_accuracy", 1) < 0.8:
        return "degraded"
    if key == "qwen3.7-flash_sorter_v13_subtype_langfuse":
        return "champion"
    return "benchmark"


def slide_sweep_full(wb: openpyxl.Workbook, records: list[dict], n: int) -> None:
    runs = _sweep_records(records)
    runs = sorted(runs, key=lambda r: r.get("timestamp", ""))
    seen: dict[str, int] = {}
    ws, row = new_slide(
        wb, "PART B · SORTER SWEEP", "Full model sweep — all sorter_v13 subtype runs",
        n,
        kicker="Every collected subtype-classification run on the champion sorter_v13 prompt (full-509, seed 42, temp 0.1, "
               "reasoning medium) — 8 distinct models incl. smokes and reruns.",
        widths=[10, 34, 16, 8, 11, 11, 11, 9, 14],
    )
    row = section_row(ws, row, "Strict / equiv / exact / fails (bootstrap CIs in the Eval Results workbook)")
    rows = []
    for r in runs:
        sc = r.get("scores", {}).get("sorter", {})
        ci = sc.get("subtype_accuracy_ci") or {}
        rows.append([
            _date(r),
            _short_run(r.get("experiment_name", ""), 30),
            _short_model(r),
            r.get("n_rows"),
            _num(sc.get("subtype_accuracy"), 4),
            _num(sc.get("subtype_accuracy_equiv"), 4),
            _num(sc.get("exact_match"), 4),
            sc.get("failure_insights", {}).get("n_failed", "—"),
            f"{_num(ci.get('lo'), 3)}–{_num(ci.get('hi'), 3)} · {_sweep_flag(r, seen)}",
        ])
    table(ws, row, 1,
          ["Date", "Run", "Model", "n", "Subtype", "Equiv", "Exact", "Fails", "CI · flag"],
          rows,
          widths=[10, 34, 16, 8, 11, 11, 11, 9, 20])
    note(ws, row, 1, "Canonical table (reconciled, 8 models): deepseek-v4-pro 0.9528 · qwen 0.9430 (champion) · gpt-4o-mini 0.9312 · "
                     "deepseek-v4-flash 0.9332 · gpt-5-nano 0.8978 · llama-3.3-70b 0.8900 · llama-4-scout 0.8880 · gpt-4.1-nano 0.8782. "
                     "Cross-model significance NOT claimed (±0.006 noise band measured on identical-prompt qwen reruns).")


def slide_sorter_champion(wb: openpyxl.Workbook, r: dict, n: int) -> None:
    ws, row = new_slide(
        wb, "PART B · SORTER SWEEP", "Champion sorter_v13 — headlines & failure modes",
        n,
        kicker="qwen3.7-flash × sorter_v13 clean full-509 rerun (the sweep baseline every model is compared against).",
        widths=[2, 24, 16, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 2],
    )
    sc = r.get("scores", {}).get("sorter", {})
    ci = sc.get("subtype_accuracy_ci") or {}
    row = section_row(ws, row, "Headlines")
    big_number(ws, row, 2, _num(sc.get("exact_match"), 4), "exact match (doc_type + subtype)")
    big_number(ws, row, 5, _num(sc.get("subtype_accuracy"), 4), f"subtype accuracy  CI [{_num(ci.get('lo'))}, {_num(ci.get('hi'))}]")
    big_number(ws, row, 8, _num(sc.get("subtype_accuracy_equiv"), 4), "subtype accuracy (equiv family)")
    big_number(ws, row, 11, str(sc.get("failure_insights", {}).get("n_failed")), "failed rows (of 509)")
    row += 3
    row = section_row(ws, row, "Failure-mode breakdown")
    modes = sc.get("failure_insights", {}).get("mode_counts") or {}
    table(ws, row, 1, ["Mode", "Count"],
          [[k, v] for k, v in modes.items()], widths=[26, 10])
    row = section_row(ws, row, "Top-confidence failure examples (with model reasoning)")
    failed = [res for res in (r.get("results") or [])
              if isinstance(res, dict) and res.get("failure_mode")]
    failed.sort(key=lambda x: x.get("confidence", 0) or 0, reverse=True)
    ex = []
    for res in failed[:4]:
        ex.append([_trunc(res.get("filename") or res.get("id") or "", 34),
                   res.get("failure_mode"),
                   _trunc((res.get("reasoning") or "")[:160], 160)])
    if ex:
        table(ws, row, 1, ["Document", "Mode", "Model reasoning"], ex, widths=[26, 16, 60])
    note(ws, row, 1, "Failure modes: family_confusion · function_over_form (doc_type miss) · equivalent_family (recovered by "
                     "equivalence) · other_fallback.")


def slide_codebook_subtypes(wb: openpyxl.Workbook, n: int) -> None:
    ws, row = new_slide(
        wb, "PART B · SORTER SWEEP", "Codebook — the 25 CUAD contract subtypes",
        n,
        kicker="agents/sorter_agent.py → CONTRACT_SUBTYPES (CUAD folder names are the ground truth).",
        widths=[2, 22, 18, 60, 16, 16, 16, 16, 16, 16, 16, 16, 16, 16, 16, 2],
    )
    row = section_row(ws, row, "Subtype key · label · definition")
    table(ws, row, 1,
          ["Key", "Label", "Definition"],
          [[k.get("key"), k.get("label"), _trunc(k.get("description", ""), 150)]
           for k in (CONTRACT_SUBTYPES or [])],
          widths=[22, 18, 90])


def slide_codebook_sorter_rules(wb: openpyxl.Workbook, fs: dict[str, Any], n: int) -> None:
    ws, row = new_slide(
        wb, "PART B · SORTER SWEEP", "Codebook — equivalences, failure modes & scoring rules",
        n,
        kicker="agents/sorter_agent.py → SUBTYPE_EQUIVALENCES · docs/SCORING.md.",
        widths=[2, 30, 60, 16, 16, 16, 16, 16, 16, 16, 16, 16, 16, 16, 16, 2],
    )
    row = section_row(ws, row, "Equivalence families (subtype_accuracy_equiv)")
    table(ws, row, 1, ["Equivalent members (share one family)", "Meaning"],
          [[", ".join(sorted(fam)), "routed defensibly within one family — counts as subtype_accuracy_equiv"]
           for fam in (SUBTYPE_EQUIVALENCES or [])],
          widths=[44, 60])
    row = section_row(ws, row, "Scoring rules (subtype surface)")
    flat = []
    if isinstance(fs, dict):
        for k, v in fs.items():
            if isinstance(v, dict):
                for k2, v2 in v.items():
                    flat.append([f"{k}.{k2}", _trunc(str(v2), 100)])
            else:
                flat.append([k, _trunc(str(v), 100)])
    table(ws, row, 1, ["Key", "Value"], flat[:20], widths=[30, 80])


# ---------------------------------------------------------------------------
# PART C — LEGALBENCH
# ---------------------------------------------------------------------------


def _lb_records(records: list[dict]) -> list[dict]:
    return [r for r in records if "legalbench" in r.get("experiment_name", "")]


def _lb_group(name: str) -> str:
    for g in ("task_v0", "task_v1", "task_v2", "task_v3", "task_v4"):
        if g in name:
            return g
    return "other"


def _lb_row(r: dict) -> list[Any]:
    s = r.get("scores", {})
    ci = s.get("exact_match_ci") or {}
    pca = s.get("per_class_accuracy") or {}
    return [
        _date(r),
        _short_run(r.get("experiment_name", ""), 36),
        _short_model(r),
        r.get("n_rows"),
        _num(s.get("exact_match"), 4),
        _num(ci.get("lo"), 3),
        _num(ci.get("hi"), 3),
        _num(pca.get("no")),
        _num(pca.get("yes")),
        _mtok(r),
        _cost(r),
    ]


def slide_lb_log(wb: openpyxl.Workbook, records: list[dict], groups: tuple[str, ...], n: int) -> None:
    runs = sorted([r for r in _lb_records(records) if _lb_group(r.get("experiment_name", "")) in groups],
                  key=lambda r: r.get("timestamp", ""))
    ws, row = new_slide(
        wb, "PART C · LEGALBENCH", "Performance log — " + ", ".join(groups),
        n,
        kicker="All collected LegalBench records for these prompt generations, chronological — exact match (binary accuracy) + "
               "bootstrap CI + per-class accuracy + tokens/cost.",
        widths=[10, 36, 14, 8, 11, 9, 9, 9, 9, 10, 10],
    )
    row = section_row(ws, row, f"{len(runs)} runs — n / exact / CI / per-class (no, yes) / tokens / est. cost")
    table(ws, row, 1,
          ["Date", "Run", "Model", "n", "Exact", "CI lo", "CI hi", "No", "Yes", "Tokens", "Cost est."],
          [_lb_row(r) for r in runs],
          widths=[10, 36, 14, 8, 11, 9, 9, 9, 9, 10, 10])


def slide_lb_summary(wb: openpyxl.Workbook, records: list[dict], n: int) -> None:
    runs = _lb_records(records)
    ws, row = new_slide(
        wb, "PART C · LEGALBENCH", "Per-task results summary",
        n,
        kicker="Best/latest exact match per (task × family) — hearsay prompt lineage (94-row test) + contract families (6-row slices).",
        widths=[2, 34, 16, 9, 11, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 2],
    )
    row = section_row(ws, row, "Heads-up summary (best exact per family)")
    best: dict[str, dict] = {}
    for r in runs:
        s = r.get("scores", {})
        em = s.get("exact_match")
        if em is None:
            continue
        name = r.get("experiment_name", "")
        family = name.split("legalbench_task_v", 1)[-1]
        family = "task_v" + family.split("_", 1)[0] if "_" in family else "task_v" + family
        if "task_v" in name:
            fam = name.replace("qwen3.7-flash_", "")
            fam = fam.split("@")[0]
            key = fam.rsplit("_", 1)[0] if fam.endswith("_sampled") else fam
        else:
            key = name
        key = key.replace("qwen3.7-flash_", "").replace("_classification_langfuse", "").replace("_usage", "").replace("_baseline", "").replace("_test_rerun94", "_test").replace("_smoke_pilot", "_test")
        if key not in best or (r.get("n_rows", 0) or 0) > (best[key]["n"] or 0):
            best[key] = {"n": r.get("n_rows"), "exact": em, "date": _date(r),
                         "ci": s.get("exact_match_ci") or {}}
    rows = [[k, best[k]["n"], _num(best[k]["exact"], 4),
             _num(best[k]["ci"].get("lo"), 3), _num(best[k]["ci"].get("hi"), 3), best[k]["date"]]
            for k in sorted(best)]
    table(ws, row, 1, ["Task / family", "n", "Exact", "CI lo", "CI hi", "Last run"],
          rows, widths=[34, 9, 11, 10, 10, 12])
    note(ws, row, 1, "hearsay task: v0 0.7766 → v1 0.8617 → v2 0.8830 on the same 94-row test (same-surface lineage). "
                     "task_v3/v4 = 7 contract families × 6-row slices (+ sampled variants); n=6 slices are diagnostic, not benchmarks.")


def slide_lb_legend(wb: openpyxl.Workbook, n: int) -> None:
    ws, row = new_slide(
        wb, "PART C · LEGALBENCH", "Task legend & scoring",
        n,
        kicker="LegalBench local mirror — data/legalbench_local/, evaluated with the repo's task-mode classification runner.",
        widths=[2, 30, 70, 16, 16, 16, 16, 16, 16, 16, 16, 16, 16, 16, 16, 2],
    )
    legend = [
        ("task_v0 / v1 / v2", "HEARSAY — LegalBench hearsay classification (94-row test, data/legalbench_local/hearsay-test.jsonl). "
                              "v0 = initial prompt; v1/v2 = prompt iterations (0.7766 → 0.8617 → 0.8830)."),
        ("task_v3", "7 contract families (binary Yes/No per document, 6-row slices): anti_assignment, audit_rights, cap_on_liability, "
                    "change_of_control, competitive_restriction_exception, covenant_not_to_sue, effective_date."),
        ("task_v4", "Same 7 families re-run (+ _sampled = seeded sample slices); family prompts refined from v3."),
        ("Scoring", "exact_match = binary accuracy (Yes/No) with percentile-bootstrap CI (seed 42, n_boot 2000); per_class_accuracy = "
                    "per-label accuracy; failure = fraction of errored rows."),
        ("prompt_mode=task", "The LegalBench runner uses --prompt-mode task + --valid-classes Yes,No with the legalbench_task_vN prompt."),
        ("Comparability", "n=6 slices are diagnostic only (tiny support); the 94-row hearsay test is the controlled same-surface lineage."),
    ]
    for k, v in legend:
        kc = ws.cell(row=row, column=2, value=k)
        kc.font = Font(bold=True, size=10)
        kc.fill = PatternFill("solid", fgColor="DEEBF7")
        kc.alignment = Alignment(horizontal="left", vertical="top", indent=1, wrap_text=True)
        kc.border = _BOX_ALL
        vc = ws.cell(row=row, column=3, value=v)
        vc.font = Font(size=10)
        vc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        vc.border = _BOX_ALL
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=14)
        ws.row_dimensions[row].height = 52.0
        row += 1
    note(ws, row + 1, 2, "Full per-record detail (params, tokens, per-class, results) lives in reports/experiment_log.jsonl / the site explorer.")


# ---------------------------------------------------------------------------
# Deck assembly
# ---------------------------------------------------------------------------


def build_deck(log_path: str, taxonomy_path: str, out_path: str) -> str:
    records = load_records_list(log_path)
    by_name: dict[str, dict] = {}
    for r in records:
        by_name[r["experiment_name"]] = r

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    slide_cover(wb)
    slide_contents(wb)

    slide_extract_lineage(wb, records, 0, 3)
    slide_extract_lineage(wb, records, 1, 4)
    champ = by_name.get(EXTRACTION_CHAMPION)
    if champ is None:
        champ = _extraction_records(records)[-1]
    slide_extract_champion_metadata(wb, champ, 5)
    slide_extract_per_field(wb, champ, 6)
    slide_extract_errors(wb, champ, 7)
    slide_extract_diagnostics(wb, champ, 8)
    fields = load_contract_fields(taxonomy_path)
    fs = load_field_scoring(taxonomy_path)
    slide_codebook_fields(wb, fields, 9)
    slide_codebook_scoring(wb, fs, 10)

    slide_sweep_full(wb, records, 11)
    sorter_champ = None
    for r in _sweep_records(records):
        if r.get("experiment_name", "").startswith("qwen3.7-flash_sorter_v13_subtype_langfuse") \
                and (r.get("scores", {}).get("sorter", {}).get("subtype_accuracy") or 0) > 0.94:
            sorter_champ = r
    slide_sorter_champion(wb, sorter_champ or _sweep_records(records)[1], 12)
    slide_codebook_subtypes(wb, 13)
    slide_codebook_sorter_rules(wb, fs, 14)

    slide_lb_log(wb, records, ("task_v0", "task_v1", "task_v2"), 15)
    slide_lb_log(wb, records, ("task_v3",), 16)
    slide_lb_log(wb, records, ("task_v4",), 17)
    slide_lb_summary(wb, records, 18)
    slide_lb_legend(wb, 19)

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)
    return out_path


def main_with_args(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--log", default=DEFAULT_LOG, help="experiment log JSONL path")
    parser.add_argument("--taxonomy", default=DEFAULT_TAXONOMY, help="taxonomy YAML path")
    parser.add_argument("--outdir", default=DEFAULT_OUTDIR, help="output directory")
    parser.add_argument("--outfile", default=DEFAULT_DECK, help="output workbook filename")
    args = parser.parse_args(argv)
    out = os.path.join(args.outdir, args.outfile)
    built = build_deck(args.log, args.taxonomy, out)
    wb = openpyxl.load_workbook(built, read_only=True)
    print(f"Wrote {built} — {len(wb.sheetnames)} slides: {', '.join(wb.sheetnames)}")
    return 0


def main() -> None:
    raise SystemExit(main_with_args(sys.argv[1:]))


if __name__ == "__main__":
    main()
