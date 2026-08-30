#!/usr/bin/env python3
"""Export a Google-Slides-style xlsx deck of the latest eval runs + codebooks.

Builds ONE workbook in ``reports/sheets/`` in which every worksheet is a
16:9 "slide" (landscape, fit-to-page, dark title banner + footer) so the
sheets read as a deck and copy-paste cleanly into Google Slides:

- **Part A — contracts specialist v32** (``..._510_full_clean`` run):
  metadata, headline scores + v31 comparison, per-field table, error
  decomposition, entity-list F1 vs raw P/R/F1, MAE/R² diagnostics, and the
  full **extraction codebook** (9 fields + types, partial-GT / containment /
  factuality rules, ambiguous band, thresholds).
- **Part B — sorter v14 subtype** (``..._sorter_v14_subtype_langfuse`` run):
  metadata + headlines (+ v13 A/B note), failure-mode breakdown + failure
  examples, per-subtype strict/equiv accuracy (derived from the row-level
  results), and the full **sorter codebook** (25 CUAD subtypes, 4 equivalence
  families, failure modes, scoring rules).
- **Part C — docclass v5 diag run** (``..._docclass_diag30b``): bonus slide
  (doc_type / subclass accuracy, per-class + per-subclass tables) and the
  sources & navigation slide.

Data is read live from ``reports/experiment_log.jsonl``,
``config/taxonomy.yaml`` and ``agents/sorter_agent.py`` — no network, no LLM.

Usage::

    python scripts/reporting/export_slides_deck.py                     # default deck
    python scripts/reporting/export_slides_deck.py --outdir reports/sheets
    python scripts/reporting/export_slides_deck.py --log reports/experiment_log.jsonl
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import datetime, timezone
from typing import Any, Callable

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

try:
    import yaml  # type: ignore

    _HAS_YAML = True
except Exception:  # pragma: no cover - yaml is an optional nicety
    _HAS_YAML = False

from agents.sorter_agent import (  # noqa: E402
    CONTRACT_SUBTYPES,
    CONTRACT_SUBTYPE_KEYS,
    SUBTYPE_EQUIVALENCES,
)

DEFAULT_LOG = "reports/experiment_log.jsonl"
DEFAULT_TAXONOMY = "config/taxonomy.yaml"
DEFAULT_OUTDIR = "reports/sheets"
DEFAULT_DECK = "contract_specialist_v32_and_sorter_v14_deck.xlsx"

# The three runs captured by this deck (experiment_name -> part).
RUN_NAMES = {
    "extraction": "qwen3.7-flash_contracts_specialist_v32_extraction_langfuse_510_full_clean",
    "sorter": "qwen3.7-flash_sorter_v14_subtype_langfuse",
    "docclass": "qwen3.7-flash_sorter_docclass_v5_docclass_diag30b",
}

MODEL_DISPLAY = {
    "qwen/qwen3.7-flash": "Qwen 3.7-Flash",
    "deepseek/deepseek-v4-flash": "DeepSeek V4 Flash",
    "deepseek/deepseek-v4-pro": "DeepSeek V4 Pro",
}


def _model(r: dict) -> str:
    return MODEL_DISPLAY.get(r.get("model", ""), r.get("model", ""))


def _get(record: dict, path: str, default: Any = None) -> Any:
    cur: Any = record
    for part in path.split("."):
        if isinstance(cur, dict) and part in cur:
            cur = cur[part]
        else:
            return default
    return cur


def _pct(v: Any) -> str:
    if v is None:
        return "—"
    try:
        return f"{float(v):.1%}"
    except (TypeError, ValueError):
        return str(v)


def _num(v: Any, nd: int = 4) -> str:
    if v is None:
        return "—"
    try:
        return f"{float(v):.{nd}f}"
    except (TypeError, ValueError):
        return str(v)


def _trunc(text: str | None, n: int = 220) -> str:
    if not text:
        return "—"
    text = " ".join(text.split())
    return text if len(text) <= n else text[: n - 1].rstrip() + "…"


# ---------------------------------------------------------------------------
# Slide geometry / styling (one worksheet = one 16:9 slide)
# ---------------------------------------------------------------------------

_NAVY = "1F4E79"
_BLUE = "2E75B6"
_BLUE_LIGHT = "DEEBF7"
_ROW_ALT = "F2F7FB"
_HILITE = "FFF2CC"
_GREEN = "E2EFDA"
_RED = "FCE4EC"
_GRAY = "808080"
_WHITE = "FFFFFF"

_THIN = Side(style="thin", color="BFBFBF")
_BOX = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_N_COLS = 16
_N_ROWS = 36
_TOP_BANNER = 46.0
_KICKER = 16.0
_SECTION_H = 20.0
_FOOTER = 16.0

_TOTAL_SLIDES = 19


def _widths(ws: openpyxl.worksheet.worksheet.Worksheet, widths: list[float]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def new_slide(
    wb: openpyxl.Workbook,
    tag: str,
    title: str,
    n: int,
    kicker: str = "",
    widths: list[float] | None = None,
) -> tuple[openpyxl.worksheet.worksheet.Worksheet, int]:
    """Create a slide sheet with the banner + footer; returns (ws, cursor_row)."""
    _short_tags = {
        "PART A · CONTRACT SPECIALIST v32": "CONTRACT v32",
        "PART B · SORTER v14": "SORTER v14",
        "PART C · DOCCLASS v5": "DOCCLASS v5",
    }
    _short_titles = {
        "Run metadata & parameters": "metadata",
        "Headline scores": "headlines",
        "Per-field scores": "per-field",
        "Error decomposition & field presence": "errors",
        "Entity lists — coverage F1 vs raw P/R/F1": "entity lists",
        "Regression diagnostics — MAE / R² / span drift": "diagnostics",
        "Fields, types & scoring class": "fields",
        "Scoring rubric & thresholds": "rubric",
        "Run metadata & headlines": "headlines",
        "Failure modes & examples": "failures",
        "Per-subtype accuracy": "per-subtype",
        "The 25 CUAD contract subtypes": "subtypes",
        "Equivalences, failure modes & scoring rules": "rules",
        "Hierarchical doc-class diag + sources": "bonus + sources",
        "Qwen lineage summary v3→v13": "qwen lineage",
        "Qwen lineage — all 30 runs": "qwen all runs",
        "Llama runs (Langfuse)": "llama runs",
    }
    name = f"{n:02d}. {_short_tags.get(tag, tag)} · {_short_titles.get(title, title)}"
    if n == 1:
        name = "01. COVER"
    ws = wb.create_sheet(name[:31])
    _widths(ws, widths or [2, 21, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 2])
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    from openpyxl.worksheet.properties import PageSetupProperties

    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins.left = ws.page_margins.right = 0.3
    ws.page_margins.top = ws.page_margins.bottom = 0.3

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=_N_COLS)
    banner = ws.cell(row=1, column=1, value=f"{tag}  ·  {title}")
    banner.font = Font(bold=True, size=15, color=_WHITE)
    banner.fill = PatternFill("solid", fgColor=_NAVY)
    banner.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = _TOP_BANNER

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=_N_COLS)
    sub = ws.cell(row=2, column=1, value=kicker)
    sub.font = Font(size=9, italic=True, color=_GRAY)
    sub.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = _KICKER

    ws.merge_cells(start_row=_N_ROWS, start_column=1, end_row=_N_ROWS, end_column=_N_COLS)
    foot = ws.cell(
        row=_N_ROWS,
        column=1,
        value=f"{DEFAULT_DECK} · slide {n}/{_TOTAL_SLIDES} · "
        f"generated {datetime.now(timezone.utc).strftime('%Y-%m-%d')} · "
        f"source: reports/experiment_log.jsonl",
    )
    foot.font = Font(size=8, color=_GRAY)
    foot.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    foot.border = Border(top=Side(style="medium", color=_GRAY))
    ws.row_dimensions[_N_ROWS].height = _FOOTER
    return ws, 4


def section_row(ws: openpyxl.worksheet.worksheet.Worksheet, row: int, text: str, col: int = 1) -> int:
    """A full-width section header bar; returns the next free row."""
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=_N_COLS)
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=True, size=11, color=_WHITE)
    cell.fill = PatternFill("solid", fgColor=_BLUE)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = _SECTION_H
    return row + 1


def table(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    row: int,
    col: int,
    headers: list[str],
    rows: list[list[Any]],
    widths: list[float] | None = None,
    fmt: Callable[[Any], Any] | None = None,
    highlight_cols: set[int] | None = None,
) -> int:
    """Write a bordered table starting at (row, col); returns the next free row.

    Rows that would spill past the slide footer (row ``_N_ROWS``) are
    dropped with a warning — slide grids have a fixed height budget.
    """
    if widths:
        for i, w in enumerate(widths, start=col):
            ws.column_dimensions[get_column_letter(i)].width = w
    for j, h in enumerate(headers):
        c = ws.cell(row=row, column=col + j, value=h)
        c.font = Font(bold=True, size=9, color=_WHITE)
        c.fill = PatternFill("solid", fgColor=_BLUE)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        c.border = _BOX
    ws.row_dimensions[row].height = _SECTION_H
    r = row + 1
    for i, data in enumerate(rows):
        if r >= _N_ROWS:
            print(f"  [warn] table truncated at slide row {_N_ROWS} "
                  f"({len(rows) - i} rows dropped)", file=sys.stderr)
            break
        for j, v in enumerate(data):
            c = ws.cell(row=r, column=col + j, value=fmt(v) if fmt else v)
            c.font = Font(size=9)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
            c.border = _BOX
            if i % 2 == 1:
                c.fill = PatternFill("solid", fgColor=_ROW_ALT)
            if highlight_cols and j in highlight_cols and v is not None:
                try:
                    if float(v) >= 0.95:
                        c.fill = PatternFill("solid", fgColor=_GREEN)
                    elif float(v) <= 0.7:
                        c.fill = PatternFill("solid", fgColor=_RED)
                except (TypeError, ValueError):
                    pass
        ws.row_dimensions[r].height = 26.0
        r += 1
    return r


def kv_table(ws: openpyxl.worksheet.worksheet.Worksheet, row: int, col: int, pairs: list[tuple[str, Any]]) -> int:
    """A two-column key/value table; returns the next free row."""
    for k, v in pairs:
        kc = ws.cell(row=row, column=col, value=k)
        kc.font = Font(bold=True, size=9)
        kc.fill = PatternFill("solid", fgColor=_BLUE_LIGHT)
        kc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        kc.border = _BOX
        vc = ws.cell(row=row, column=col + 1, value=v)
        vc.font = Font(size=9)
        vc.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        vc.border = _BOX
        ws.row_dimensions[row].height = 24.0
        row += 1
    return row


def big_number(ws: openpyxl.worksheet.worksheet.Worksheet, row: int, col: int, value: Any, label: str) -> None:
    """A large callout value + caption (Google-Slides stat-card feel)."""
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=True, size=20, color=_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor=_HILITE)
    c.border = _BOX
    ws.row_dimensions[row].height = 34.0
    l = ws.cell(row=row + 1, column=col, value=label)
    l.font = Font(size=8, italic=True, color=_GRAY)
    l.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    ws.row_dimensions[row + 1].height = 22.0


def note(ws: openpyxl.worksheet.worksheet.Worksheet, row: int, col: int, text: str, italic: bool = True) -> int:
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=_N_COLS - 1)
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(size=8, italic=italic, color=_GRAY)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    ws.row_dimensions[row].height = 26.0
    return row + 1


# ---------------------------------------------------------------------------
# Data loaders
# ---------------------------------------------------------------------------


def load_records(log_path: str) -> dict[str, dict]:
    """Index experiment-log records by experiment name (last record wins)."""
    records: dict[str, dict] = {}
    with open(log_path, encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line:
                continue
            r = json.loads(line)
            records[r["experiment_name"]] = r
    return records


def load_records_list(log_path: str) -> list[dict]:
    """Every experiment-log record in file order (no dedup)."""
    records: list[dict] = []
    with open(log_path, encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line:
                continue
            records.append(json.loads(line))
    return records


def load_contract_fields(taxonomy_path: str) -> dict[str, str]:
    """The contract doc-class field_types from config/taxonomy.yaml."""
    if not _HAS_YAML:
        return {}
    with open(taxonomy_path, encoding="utf-8") as fh:
        tax = yaml.safe_load(fh)
    for cls in tax.get("doc_classes", []):
        if cls.get("key") == "contract":
            return dict(cls.get("field_types", {}))
    return {}


def load_field_scoring(taxonomy_path: str) -> dict[str, Any]:
    if not _HAS_YAML:
        return {}
    with open(taxonomy_path, encoding="utf-8") as fh:
        return yaml.safe_load(fh).get("field_scoring", {})


# ---------------------------------------------------------------------------
# Slide builders
# ---------------------------------------------------------------------------


def slide_cover(wb: openpyxl.Workbook) -> None:
    ws, _ = new_slide(wb, "COVER", "Eval results + codebooks", 1, widths=[2, 26, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 2])
    row = 8
    c = ws.cell(row=row, column=3, value="Contract Specialist v32  ·  Sorter v14  ·  Docclass v5")
    c.font = Font(bold=True, size=24, color=_NAVY)
    row += 2
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=13)
    c = ws.cell(row=row, column=3, value="llm-entity-extraction — most recent run results and full task codebooks, as a slide deck")
    c.font = Font(size=13, italic=True, color=_GRAY)
    row += 3
    sub = [
        "Part A — Contracts specialist v32 · full-corpus 509-doc chunked extraction (2026-08-16)",
        "Part B — Sorter v14 · full-corpus 509-doc contract-subtype classification (2026-08-16)",
        "Part C — Docclass v5 · 30-doc hierarchical doc-class diagnostic (2026-08-16)",
        "Every sheet is a 16:9 slide (landscape, fit-to-page) — import into Google Slides by copying each sheet.",
    ]
    for s in sub:
        c = ws.cell(row=row, column=3, value="•  " + s)
        c.font = Font(size=11)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[row].height = 22.0
        row += 1
    row += 2
    c = ws.cell(row=row, column=3, value="Sources: reports/experiment_log.jsonl · config/taxonomy.yaml · agents/sorter_agent.py · docs/SCORING.md")
    c.font = Font(size=9, italic=True, color=_GRAY)


def slide_contents(wb: openpyxl.Workbook) -> None:
    ws, row = new_slide(wb, "CONTENTS", "Deck map", 2)
    row = section_row(ws, row, "Part A — Contracts specialist v32 (extraction, n=509, chunked)")
    items = [
        "03 · Run metadata & parameters · 04 · Headline scores + v31 comparison · 05 · Per-field scores",
        "06 · Error decomposition (exact / partial / miss) · 07 · Entity lists (F1 vs raw P/R/F1) · 08 · MAE / R² diagnostics",
        "09 · Codebook: extraction fields & types · 10 · Codebook: scoring rubric & thresholds",
    ]
    for s in items:
        c = ws.cell(row=row, column=2, value="•  " + s)
        c.font = Font(size=10)
        ws.row_dimensions[row].height = 20.0
        row += 1
    row += 1
    row = section_row(ws, row, "Part B — Sorter (contract-subtype classification, n=509)")
    for s in [
        "11 · Sorter v14 run metadata & headlines (+ v13 A/B note) · 12 · Failure modes & examples · 13 · Per-subtype accuracy",
        "14 · Qwen lineage summary v3→v13 · 15 · Qwen lineage — all 30 runs · 16 · Llama runs (Langfuse)",
        "17 · Codebook: 25 CUAD subtypes · 18 · Codebook: equivalence families, failure modes, scoring rules",
    ]:
        c = ws.cell(row=row, column=2, value="•  " + s)
        c.font = Font(size=10)
        ws.row_dimensions[row].height = 20.0
        row += 1
    row += 1
    row = section_row(ws, row, "Part C — Docclass + sources")
    for s in ["19 · Docclass v5 diag-30 bonus slide + sources & navigation"]:
        c = ws.cell(row=row, column=2, value="•  " + s)
        c.font = Font(size=10)
        ws.row_dimensions[row].height = 20.0
        row += 1
    note(ws, row + 2, 2, "Same-surface comparisons only: the 509-doc runs are the full corpus (seed 42); the docclass slide is a 30-doc diagnostic, not comparable to 509-doc numbers.")


def slide_extract_metadata(wb: openpyxl.Workbook, r: dict) -> None:
    ws, row = new_slide(
        wb, "PART A · CONTRACT SPECIALIST v32", "Run metadata & parameters", 3,
        kicker=f"{_model(r)} × contracts_specialist_v32 — full-corpus chunked extraction (510-full-clean)",
    )
    p = r.get("parameters", {})
    ds = r.get("data_source", {})
    t = r.get("tokens", {})
    row = section_row(ws, row, "Run identity")
    row = kv_table(ws, row, 2, [
        ("Experiment", r.get("experiment_name")),
        ("Task", r.get("task")),
        ("Model", f"{_model(r)} ({r.get('model')})"),
        ("Prompt version", r.get("prompt_version")),
        ("Timestamp", r.get("timestamp", "")[:19].replace("T", " ")),
        ("Git snapshot", f"{r.get('git', {}).get('commit', '?')}" + (" (dirty tree)" if r.get('git', {}).get('dirty') else "")),
    ])
    row = section_row(ws, row, "Dataset")
    row = kv_table(ws, row, 2, [
        ("Project", ds.get("project")),
        ("Ground truth", f"{ds.get('ground_truth')} ({ds.get('ground_truth_mode')})"),
        ("Dataset fingerprint", ds.get("dataset_fingerprint", "")[:16] + "…"),
        ("n samples", ds.get("n_samples")),
        ("Seed", ds.get("seed")),
        ("Master labels", str(ds.get("master_labels", "—")).rsplit("/", 1)[-1]),
    ])
    row = section_row(ws, row, "Scoring / tracing parameters")
    row = kv_table(ws, row, 2, [
        ("Temperature / max tokens", f"{p.get('temperature')} / {p.get('max_tokens')}"),
        ("Reasoning effort", p.get("reasoning_effort")),
        ("Max input chars", p.get("max_input_chars")),
        ("Chunking", f"chunked: {p.get('chunk_chars')} chars, {p.get('chunk_overlap')} overlap"),
        ("Manifest", str(p.get("manifest", "—")).rsplit("/", 1)[-1]),
        ("Tracing backend", f"{p.get('tracing_backend')} → {p.get('tracing', {}).get('project')}"),
    ])
    row = section_row(ws, row, "Tokens & cost")
    row = kv_table(ws, row, 2, [
        ("Prompt tokens", f"{t.get('prompt_tokens', 0):,}"),
        ("Completion tokens", f"{t.get('completion_tokens', 0):,}"),
        ("Total tokens", f"{t.get('total_tokens', 0):,}"),
        ("Est. cost (local price table)", f"${t.get('cost_estimated_usd', 0):.4f}"),
        ("Rows with usage", t.get("rows_with_usage")),
    ])


def slide_extract_headlines(wb: openpyxl.Workbook, r: dict) -> None:
    ws, row = new_slide(
        wb, "PART A · CONTRACT SPECIALIST v32", "Headline scores", 4,
        kicker="Deterministic field-type-aware scoring — 500 scored rows (509 sampled), seed 42, percentile bootstrap n=2000",
    )
    s = r["scores"]
    ci = s.get("overall_extraction_score_ci", {})
    big_number(ws, row, 2, _num(s.get("overall_extraction_score")), "overall extraction score")
    big_number(ws, row, 4, f"[{_num(ci.get('lo'))}, {_num(ci.get('hi'))}]", "95% CI (half-width " + _num(ci.get("half")) + ")")
    big_number(ws, row, 6, _pct(s.get("field_presence")), "field presence (fields with any GT)")
    big_number(ws, row, 8, _pct(s.get("schema_valid")), "schema-valid rows")
    big_number(ws, row, 10, _pct(s.get("overall_verified_precision")), "verified precision (factuality guard)")
    big_number(ws, row, 12, _pct(s.get("category_presence")), "CUAD category presence")
    row += 4
    row = section_row(ws, row, "Champion comparison — v32 vs v31 (same 509-doc surface, seed 42, chunked)")
    row = table(ws, row, 2,
                ["Prompt", "Overall", "Field presence", "Schema valid", "Verified precision", "Category presence", "Effective-date field"],
                [
                    ["contracts_specialist_v31 (510-full)", _num(s31 := 0.8737), _pct(0.9655), _pct(1.0), _pct(0.9799), _pct(0.8474), _num(0.8577)],
                    ["contracts_specialist_v32 (510-full-clean)", _num(s.get("overall_extraction_score")), _pct(s.get("field_presence")), _pct(s.get("schema_valid")), _pct(s.get("overall_verified_precision")), _pct(s.get("category_presence")), _num(s.get("per_field", {}).get("effective_date"))],
                    ["Δ (v32 − v31)", f"{s.get('overall_extraction_score') - s31:+.4f}", "", "", "", "", f"{s.get('per_field', {}).get('effective_date') - 0.8577:+.4f}"],
                ],
                highlight_cols={1, 2, 3, 4, 5, 6})
    row = note(ws, row, 2, "Verdict (memo contracts_specialist_v32.md): Δ +0.0053 on the 495-row intersection, bootstrap CI [−0.0052, +0.0159], P(Δ≤0)=0.1715 — INSIDE the ±0.011 noise band → logic repair, NOT a claimed win; v31 stays aggregate champion, v32 = effective_date field specialist (+0.0171 field).")
    note(ws, row, 2, "v31 numbers are the 510-full run (2026-08-15) — same surface, same scorer; the v32 clean rerun supersedes the earlier 510_full run (52 transient errors).")


def slide_extract_per_field(wb: openpyxl.Workbook, r: dict) -> None:
    ws, row = new_slide(
        wb, "PART A · CONTRACT SPECIALIST v32", "Per-field scores", 5,
        kicker="score = type-aware deterministic scoring; verified precision = factuality-guarded; hallucination rate = ungrounded predicted items",
    )
    s = r["scores"]
    per = s.get("per_field", {})
    vp = s.get("verified_precision", {})
    hr = s.get("hallucination_rate", {})
    rows = [
        [f, _num(per.get(f)), _pct(vp.get(f)), _pct(hr.get(f))]
        for f in ["document_name", "parties", "effective_date", "governing_law", "term_length", "renewal_terms", "key_obligations", "termination_clauses"]
    ]
    rows.append(["MEAN", _num(sum(per.values()) / len(per)), _pct(sum(vp.values()) / len(vp)), _pct(sum(hr.values()) / len(hr))])
    row = table(ws, row, 2, ["Field", "Score", "Verified precision", "Hallucination rate"], rows, widths=[2, 24, 20, 24, 24, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 2], highlight_cols={2, 3, 4})
    row = note(ws, row, 2, "Strongest: document_name 0.9869 / termination_clauses 0.8933 · Weakest: key_obligations 0.7881 (partial-GT coverage metric), renewal_terms 0.7921, term_length 0.8083 (containment fields).")
    note(ws, row, 2, "List fields (parties / key_obligations / termination_clauses) are scored by GROUND-TRUTH COVERAGE — recall over matched CUAD labels, not F1 (labels are partial samples; the model is usually more complete).")


def slide_extract_errors(wb: openpyxl.Workbook, r: dict) -> None:
    ws, row = new_slide(
        wb, "PART A · CONTRACT SPECIALIST v32", "Error decomposition & field presence", 6,
        kicker="per-field exact / partial / miss rates (3,121 fields scored) — the error budget by field",
    )
    d = r["scores"]["diagnostics"]
    ed = d.get("error_decomposition", {})
    fp = d.get("field_presence_per_field", {})
    rows = []
    for f in ["document_name", "parties", "effective_date", "term_length", "termination_clauses", "governing_law", "key_obligations", "renewal_terms"]:
        e = ed.get(f, {})
        rows.append([f, _pct(e.get("exact_rate")), _pct(e.get("partial_rate")), _pct(e.get("miss_rate")), _pct(fp.get(f))])
    rows.append(["OVERALL", _pct(d.get("field_exact_rate")), _pct(d.get("field_partial_rate")), _pct(d.get("field_miss_rate")), _pct(s := r["scores"].get("field_presence"))])
    row = table(ws, row, 2, ["Field", "Exact", "Partial", "Miss", "Field presence"], rows, widths=[2, 24, 12, 12, 12, 16, 11, 11, 11, 11, 11, 11, 11, 11, 11, 2], highlight_cols={2, 3, 4, 5})
    row = note(ws, row, 2, "key_obligations is the partial-score workhorse: 64.4% partial (correct spans within family sections, near misses at sentence level) — expected for a partial-GT extraction surface, not a miss.")
    note(ws, row, 2, "contract_value presence 0.40 / renewal_terms 0.374 (only scored where CUAD GT exists; 0.4-0.6 of docs have such clauses).")


def slide_extract_entity_lists(wb: openpyxl.Workbook, r: dict) -> None:
    ws, row = new_slide(
        wb, "PART A · CONTRACT SPECIALIST v32", "Entity lists — coverage F1 vs raw P/R/F1", 7,
        kicker="partial-GT list fields: reported F1 = GT-coverage recall; raw precision/recall = exact-item statistics (always the harsher view)",
    )
    s = r["scores"]
    d = s["diagnostics"]
    f1 = s.get("entity_list_f1", {})
    lp = d.get("entity_list_precision", {})
    lr = d.get("entity_list_recall", {})
    lf = d.get("entity_list_raw_f1", {})
    rows = [
        [f, _num(f1.get(f)), _num(lr.get(f)), _num(lp.get(f)), _num(lf.get(f))]
        for f in ["parties", "key_obligations", "termination_clauses"]
    ]
    row = table(ws, row, 2, ["Field", "Score (GT coverage)", "Raw recall", "Raw precision", "Raw F1"], rows, widths=[2, 24, 20, 14, 16, 16, 11, 11, 11, 11, 11, 11, 11, 11, 11, 2], highlight_cols={2, 3, 4, 5})
    row += 1
    row = section_row(ws, row, "Micro-aggregates (all list items)")
    row = kv_table(ws, row, 2, [
        ("List items predicted", f"{d.get('list_micro_n_predicted'):,}  ·  expected {d.get('list_micro_n_expected'):,}"),
        ("Micro precision / recall / F1", f"{_num(d.get('list_micro_precision'))} / {_num(d.get('list_micro_recall'))} / {_num(d.get('list_micro_f1'))}"),
        ("Matched items", d.get("list_micro_matched")),
    ])
    note(ws, row, 2, "Why raw precision is low: the model returns 2–12 clean items per doc while CUAD labels sample only some clause answers — unmatched predicted items are correct-but-unlabeled (verified_precision 0.99 shows they ARE grounded in the source).")
    note(ws, row, 2, "Bipartite matching (Hungarian) over pairwise similarity, threshold 0.6; role-word labels (e.g. 'Shipper.') match whenever any party is named.")


def slide_extract_diagnostics(wb: openpyxl.Workbook, r: dict) -> None:
    ws, row = new_slide(
        wb, "PART A · CONTRACT SPECIALIST v32", "Regression diagnostics — MAE / R² / span drift", 8,
        kicker="post-hoc run-level diagnostics (src/metrics.py) — ALWAYS read support sizes with the numbers",
    )
    d = r["scores"]["diagnostics"]
    row = table(ws, row, 2,
                ["Metric", "MAE", "Median abs. err", "R²", "Pairs"],
                [
                    ["Dates (days)", _num(d.get("date_mae_days"), 1), _num(d.get("date_median_ae_days"), 1), _num(d.get("date_r2")), d.get("date_n_pairs")],
                    ["Durations (days)", _num(d.get("duration_mae_days"), 1), _num(d.get("duration_median_ae_days"), 1), _num(d.get("duration_r2")), d.get("duration_n_pairs")],
                    ["Money (USD)", "—", "—", "—", d.get("money_n_pairs", 0)],
                    ["Span count (per doc)", _num(d.get("span_count_mae"), 2), "—", "—", d.get("span_count_n_docs")],
                ],
                widths=[2, 26, 12, 14, 10, 10, 11, 11, 11, 11, 11, 11, 11, 11, 11, 2])
    row += 1
    row = section_row(ws, row, "Per-field date MAE + span-count drift")
    row = table(ws, row, 2,
                ["Field", "Date MAE (days)", "Span-count MAE", "Signed mean"],
                [
                    ["effective_date", _num(d.get("date_mae_per_field", {}).get("effective_date", {}), 1), "—", "—"],
                    ["term_length", _num(d.get("date_mae_per_field", {}).get("term_length", {}), 1), "—", "—"],
                    ["key_obligations (list)", "—", _num(d.get("span_count_mae_per_field", {}).get("key_obligations", {}), 2), _num(d.get("span_count_signed_mean_per_field", {}).get("key_obligations", {}), 2)],
                    ["termination_clauses (list)", "—", _num(d.get("span_count_mae_per_field", {}).get("termination_clauses", {}), 2), _num(d.get("span_count_signed_mean_per_field", {}).get("termination_clauses", {}), 2)],
                    ["Span count overall", "—", _num(d.get("span_count_mae"), 2), _num(d.get("span_count_signed_mean"), 2)],
                ],
                widths=[2, 26, 18, 16, 14, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 2])
    note(ws, row, 2, f"Signed mean +{_num(d.get('span_count_signed_mean'), 2)} → model over-extracts ~5 spans/doc on average vs GT clause lists (consistent with partial-GT semantics).")
    note(ws, row, 2, "R² = 1 − SS_res/SS_tot (negatives kept); money pairs 0 on this corpus (no currency GT values in the partial labels); duration MAE 423.9d is driven by a long tail of multi-year terms — read median AE with it.")


def slide_codebook_fields(wb: openpyxl.Workbook, fields: dict[str, str]) -> None:
    ws, row = new_slide(
        wb, "CODEBOOK · EXTRACTION", "Fields, types & scoring class", 9,
        kicker="contract doc-class schema (config/taxonomy.yaml) — field_types drive the deterministic scorer",
    )
    partial = {"parties", "key_obligations", "termination_clauses"}
    containment = {"governing_law", "term_length", "renewal_terms"}
    types = {
        "name": "exact-after-normalize / fuzzy name match",
        "date": "ISO-normalized date (exact after normalize)",
        "money": "currency amount (plain number, no commas/symbols)",
        "free_text": "token F1 vs expected (containment for clause fields)",
        "entity_list:name": "entity list — bipartite match over names",
        "entity_list:free_text": "entity list — bipartite match over clause text",
    }
    rows = []
    for f, t in fields.items():
        flag = "PARTIAL-GT" if f in partial else ("CONTAINMENT" if f in containment else "scalar")
        rows.append([f, t, flag, types.get(t, "—")])
    row = table(ws, row, 2, ["Field", "Type", "Scoring class", "Scoring rule"], rows, widths=[2, 22, 26, 16, 40, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 2])
    row = note(ws, row, 2, "PARTIAL-GT = labels are a partial sample of the document → scored by GT-coverage recall (not F1). CONTAINMENT = expected-within-predicted token containment (returning the sentence + riders no longer zeroes the field).")
    note(ws, row, 2, "v24 format discipline: ISO dates, leading duration phrases, plain currency amounts — keeps the MAE/R² pair counts alive. key_obligations scoped to CUAD restriction/covenant families (mean 7.4, max 22 items).")


def slide_codebook_scoring(wb: openpyxl.Workbook, fs: dict[str, Any]) -> None:
    ws, row = new_slide(
        wb, "CODEBOOK · EXTRACTION", "Scoring rubric & thresholds", 10,
        kicker="field-type-aware deterministic scoring (docs/SCORING.md §3; config/taxonomy.yaml field_scoring)",
    )
    row = table(ws, row, 2,
                ["Rule", "Setting"],
                [
                    ["Bipartite match threshold (entity lists)", str(fs.get("bipartite_match_threshold", 0.6))],
                    ["Ambiguous band (judge escalation gate)", str(fs.get("ambiguous_band"))],
                    ["Embedding rescue trigger", f"string score < {fs.get('embedding_rescue_below', 0.7)} → all-MiniLM-L6-v2 (local) / OpenRouter fallback"],
                    ["Factuality: token coverage for grounding", str(fs.get("factuality_verification", {}).get("token_coverage", 0.7))],
                    ["Partial-GT fields (coverage scoring)", ", ".join(fs.get("partial_gt_fields", []))],
                    ["Containment fields", ", ".join(fs.get("containment_fields", []))],
                ],
                widths=[2, 40, 58, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 2])
    row += 1
    row = section_row(ws, row, "Field-type scoring classes")
    row = table(ws, row, 2,
                ["Type", "Score"],
                [
                    ["date", "exact after ISO normalization (MAE in days; R² across docs)"],
                    ["money", "plain-amount exact / MAE (USD); pair counts with support"],
                    ["id", "exact after normalization"],
                    ["name", "fuzzy match (normalized token overlap)"],
                    ["free_text", "token F1; containment fields use expected-within-predicted tokens"],
                    ["entity_list:*", "Hungarian bipartite matching over pairwise similarity ≥ threshold; partial-GT lists scored by GT coverage"],
                ],
                widths=[2, 24, 76, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 2])
    row = note(ws, row, 2, "Factuality guard: every predicted list item must match a GT label OR be grounded in the source document (token coverage ≥ 0.7); neither → hallucination → drives verified_precision down. Tracker consistency: per-field score = *_f1 tracker = overall_extraction_score (same list score everywhere).")
    note(ws, row, 2, "v24+ reasoning trace: REQUIRED per-field reasoning (summary + entries[{field, evidence, section_ref}]) emitted BEFORE the extraction JSON, unioned across chunk windows, NEVER scored.")


def slide_sorter_headlines(wb: openpyxl.Workbook, r: dict) -> None:
    ws, row = new_slide(
        wb, "PART B · SORTER v14", "Run metadata & headlines", 11,
        kicker=f"{_model(r)} × sorter_v14 — full-corpus 509-doc contract-subtype classification",
    )
    p = r.get("parameters", {})
    t = r.get("tokens", {})
    sc = r["scores"]["sorter"]
    ci = sc.get("subtype_accuracy_ci", {})
    big_number(ws, row, 2, _num(sc.get("exact_match")), "exact match (doc_type + subtype)")
    big_number(ws, row, 4, _num(sc.get("subtype_accuracy")), f"subtype accuracy  CI [{_num(ci.get('lo'))}, {_num(ci.get('hi'))}]")
    big_number(ws, row, 6, _num(sc.get("subtype_accuracy_equiv")), "subtype accuracy (equivalent family)")
    big_number(ws, row, 8, _num(sc.get("confidence")), "mean model confidence")
    big_number(ws, row, 10, f"{sc.get('failure_insights', {}).get('n_failed')}/509", "failed rows")
    row += 4
    row = section_row(ws, row, "Parameters & cost")
    row = kv_table(ws, row, 2, [
        ("Prompt version / task", f"sorter_v14 · {r.get('task')}"),
        ("Dataset", f"{r.get('data_source', {}).get('project')} · fp {str(r.get('data_source', {}).get('dataset_fingerprint', ''))[:16]}… · n={r.get('data_source', {}).get('n_samples')} · seed {r.get('data_source', {}).get('seed')}"),
        ("Temp / max tokens / reasoning", f"{p.get('temperature')} / {p.get('max_tokens')} / {p.get('reasoning_effort')} (medium — 25 near-synonymous families)"),
        ("Tokens / est. cost", f"{t.get('total', {}).get('total_tokens', 0):,} · ${t.get('total', {}).get('cost_estimated_usd', 0):.4f}"),
        ("Manifest / tracing", f"{str(p.get('manifest', '—')).rsplit('/', 1)[-1]} · {p.get('tracing_backend')}"),
        ("Git snapshot", f"{r.get('git', {}).get('commit', '?')}" + (" (dirty tree)" if r.get('git', {}).get('dirty') else "")),
    ])
    row = section_row(ws, row, "Champion comparison — v14 vs v13 (same surface, seed 42, temp 0.1, reasoning medium)")
    row = table(ws, row, 2,
                ["Prompt", "Exact", "Subtype strict", "Subtype equiv", "Failed"],
                [
                    ["sorter_v13 (clean 509)", _num(0.9961), _num(0.9430), _num(0.9470), "29"],
                    ["sorter_v14 (509)", _num(sc.get("exact_match")), _num(sc.get("subtype_accuracy")), _num(sc.get("subtype_accuracy_equiv")), str(sc.get("failure_insights", {}).get("n_failed"))],
                    ["Δ (v14 − v13)", _num(0.0), _num(sc.get("subtype_accuracy") - 0.9430, 4), _num(sc.get("subtype_accuracy_equiv") - 0.9470, 4), ""],
                ],
                widths=[2, 26, 16, 18, 18, 14, 11, 11, 11, 11, 11, 11, 11, 11, 11, 2],
                highlight_cols={2, 3, 4})
    note(ws, row, 2, "Verdict (memo sorter_v14.md): Δ −0.0059, paired CI [−0.0177, +0.0059], P(Δ≤0)=0.8765 — INSIDE the ±0.006 noise band, negative direction → NOT a claimed win; v13 stays champion. Rule 30 DID recover Audible + PACIRA (marketing cell 14/17 → 16/17).")


def slide_sorter_failures(wb: openpyxl.Workbook, r: dict) -> None:
    ws, row = new_slide(
        wb, "PART B · SORTER v14", "Failure modes & examples", 12,
        kicker="failure_insights — mode counts + per-failed-row {expected, predicted, confidence, reasoning}",
    )
    fi = r["scores"]["sorter"]["failure_insights"]
    mc = fi.get("mode_counts", {})
    row = table(ws, row, 2,
                ["Failure mode", "Count", "Meaning"],
                [
                    ["family_confusion", mc.get("family_confusion", 0), "correct family tier, wrong subtype family (no equivalence recover)"],
                    ["equivalent_family", mc.get("equivalent_family", 0), "recovered by SUBTYPE_EQUIVALENCES (counted in equiv accuracy)"],
                    ["function_over_form", mc.get("function_over_form", 0), "doc_type-level miss (non-contract vs contract)"],
                    ["other_fallback", mc.get("other_fallback", 0), "model fell back to a generic family"],
                ],
                widths=[2, 24, 10, 66, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 2])
    row += 1
    row = section_row(ws, row, "Highest-confidence failures (predicted but wrong)")
    fails = sorted([f for f in fi.get("failures", []) if f.get("predicted")], key=lambda f: -(f.get("confidence") or 0))[:3]
    hdr = ["Doc (short)", "Expected", "Predicted", "Confidence", "Mode", "Model reasoning (excerpt)"]
    rows = [
        [f.get("filename", "")[:48], f.get("expected"), f.get("predicted"), _pct(f.get("confidence")), f.get("mode"), _trunc(f.get("reasoning"), 180)]
        for f in fails
    ]
    row = table(ws, row, 2, hdr, rows, widths=[2, 26, 12, 12, 10, 14, 34, 11, 11, 11, 11, 11, 11, 11, 11, 2])
    note(ws, row, 2, "Pattern: title-substance conflicts (escrow attached to a license; 'sponsorship' naming an agency relationship) — the title-wins rules (23/24/26/28/30) deliberate on exactly this boundary.")
    note(ws, row, 2, "Equivalence recovers 2 (distributor, license) — the families interchangeable by contract (see codebook slide 15).")


def slide_sorter_per_subtype(wb: openpyxl.Workbook, r: dict) -> None:
    ws, row = new_slide(
        wb, "PART B · SORTER v14", "Per-subtype accuracy", 13,
        kicker="derived from the 509 row-level results — strict (subtype_ok) and equivalent-family (subtype_ok_equiv) accuracy by CUAD family",
    )
    acc: dict[str, list[int]] = {}
    for rr in r["results"]:
        s = rr.get("sorter", {})
        exp = s.get("expected_subtype")
        if exp is None:
            continue
        a = acc.setdefault(exp, [0, 0, 0])
        a[0] += 1
        a[1] += int(bool(s.get("subtype_ok")))
        a[2] += int(bool(s.get("subtype_ok_equiv")))
    order = CONTRACT_SUBTYPE_KEYS
    left, right = order[:13], order[13:]

    def row_for(k: str) -> list[Any]:
        t, c, q = acc[k]
        return [k, t, _num(c / t), _num(q / t)]

    rows_l = [row_for(k) for k in left if k in acc]
    rows_r = [row_for(k) for k in right if k in acc]
    tot = [sum(acc[k][i] for k in acc) for i in range(3)]
    rows_l.append(["TOTAL", tot[0], _num(tot[1] / tot[0]), _num(tot[2] / tot[0])])
    rows_r.append(["", "", "", ""])
    hdr = ["Subtype", "n", "Strict", "Equiv"]
    r1 = table(ws, row, 2, hdr, rows_l, widths=[2, 26, 8, 10, 10, 14, 11, 11, 11, 11, 11, 11, 11, 11, 11, 2], highlight_cols={3, 4})
    r2 = table(ws, row, 9, hdr, rows_r, highlight_cols={3, 4})
    row = max(r1, r2)
    note(ws, row, 2, "Best families (1.000): consulting, endorsement, ip, joint_venture, maintenance, manufacturing, non_compete_no_solicit, promotion, strategic_alliance, transportation · Weakest: development 0.750, reseller 0.833, supply 0.833, outsourcing 0.833, service 0.857.")
    note(ws, row, 2, "Development 0.750 = 7 misses — the development↔license boundary is the single largest confusion cluster (equiv recovers 1).")


_VERSION_NOTES = {
    "sorter_v3": "early lineage baseline (50/195-doc surfaces)",
    "sorter_v4": "—",
    "sorter_v5": "first full-509 run",
    "sorter_v6": "0.9436 @195 (ab200 arm)",
    "sorter_v7": "+0.82pp @250 (KANBAN-003)",
    "sorter_v8": "+2.06pp @243 (promotion cluster fixed)",
    "sorter_v9": "former aggregate champion (KANBAN-012)",
    "sorter_v10": "marketing title-wins — logic repair (KANBAN-013)",
    "sorter_v11": "affiliate-is-not-marketing — logic repair (KANBAN-013)",
    "sorter_v12": "strategic_alliance title-wins — logic repair (KANBAN-023)",
    "sorter_v13": "AGGREGATE CHAMPION — maintenance title-wins (KANBAN-031)",
}


def _qwen_subtype_runs(records_list: list[dict]) -> list[dict]:
    """All qwen3.7-flash sorter_v3..v13 subtype runs, chronological.

    Takes the RAW record LIST (not the deduped name->record dict) — several
    versions reran under the same experiment name (v3 ×4, v6 ×3, v9 ×4,
    v13 ×2) and deduping would drop runs.
    """
    out = []
    for r in records_list:
        if r.get("task") != "subtype_classification":
            continue
        if (r.get("model") or "") != "qwen/qwen3.7-flash":
            continue
        pv = (r.get("prompt_versions") or {}).get("sorter")
        if not pv or not pv.startswith("sorter_v"):
            continue
        try:
            num = int(pv.rsplit("_v", 1)[1])
        except ValueError:
            continue
        if 3 <= num <= 13:
            out.append(r)
    out.sort(key=lambda r: r.get("timestamp", ""))
    return out


def slide_qwen_lineage_summary(wb: openpyxl.Workbook, records_list: list[dict]) -> None:
    ws, row = new_slide(
        wb, "PART B · SORTER", "Qwen lineage summary v3→v13", 14,
        kicker="qwen3.7-flash × sorter prompt lineage — best full-surface run per version (max n, then max strict)",
    )
    runs = _qwen_subtype_runs(records_list)
    by_ver: dict[str, list[dict]] = {}
    for r in runs:
        by_ver.setdefault((r.get("prompt_versions") or {}).get("sorter"), []).append(r)
    rows = []
    for ver in [f"sorter_v{i}" for i in range(3, 14)]:
        cands = sorted(by_ver.get(ver, []),
                       key=lambda r: (r.get("data_source", {}).get("n_samples") or 0,
                                      r["scores"]["sorter"].get("subtype_accuracy") or 0,
                                      r.get("timestamp", "")))
        if not cands:
            continue
        best = cands[-1]
        sc = best["scores"]["sorter"]
        short = best["experiment_name"].replace("qwen3.7-flash_", "").replace("_subtype_langfuse", "").replace("_subtype", "").replace("_rerun_509_clean", "").replace("_ab200", "").replace("_rerun", "")
        rows.append([ver, short, best.get("data_source", {}).get("n_samples"),
                     _num(sc.get("subtype_accuracy")), _num(sc.get("exact_match")),
                     _VERSION_NOTES.get(ver, "")])
    row = table(ws, row, 2, ["Version", "Reference run", "n", "Strict", "Exact", "Notes"], rows,
                widths=[2, 15, 40, 7, 10, 10, 44, 11, 11, 11, 11, 11, 11, 11, 11, 2],
                highlight_cols={4, 5})
    note(ws, row, 2, "Reference-run rule: for each version, the run on the LARGEST surface (509 > 243 > 195 > 50); ties broken by strict accuracy then recency — degraded reruns (0.0000 / connection-error runs) never surface here.")
    note(ws, row, 2, "The 509-doc surfaces are the only directly comparable ones (same seed 42): v5 0.8585 → v6 0.9312 → v8 0.9018 → v9 0.9175 → v12 0.9293 → v13 0.9430. 243-doc runs (v7/v10/v11) are a different surface — never compare across surfaces.")


def slide_qwen_lineage_runs(wb: openpyxl.Workbook, records_list: list[dict]) -> None:
    ws, row = new_slide(
        wb, "PART B · SORTER", "Qwen lineage — all 30 runs", 15,
        kicker="all 30 qwen3.7-flash sorter_v3→v13 subtype runs in the experiment log, chronological (both v13 qwen rows: the degraded first run and the clean champion rerun)",
    )
    runs = _qwen_subtype_runs(records_list)
    rows = [[r.get("timestamp", "")[5:16].replace("T", " "), r.get("experiment_name", "").replace("qwen3.7-flash_", ""),
             r.get("data_source", {}).get("n_samples"), _num(r["scores"]["sorter"].get("subtype_accuracy")),
             _num(r["scores"]["sorter"].get("exact_match"))] for r in runs]
    half = (len(rows) + 1) // 2
    hdr = ["Date", "Run", "n", "Strict", "Exact"]
    r1 = table(ws, row, 2, hdr, rows[:half], widths=[2, 13, 30, 6, 9, 9, 10, 11, 11, 11, 11, 11, 11, 11, 11, 2],
               highlight_cols={4, 5})
    r2 = table(ws, row, 8, hdr, rows[half:], highlight_cols={4, 5})
    row = max(r1, r2)
    note(ws, row, 2, "Degraded runs kept for truthfulness: v11 first run (0.0000 — all rows errored) and v13 first run (0.7741 — 93 connection errors, KANBAN-031); each was superseded by a clean rerun within the hour.")
    note(ws, row, 2, "Surfaces: n=509 (full corpus, seed 42) · n=243 (stratified) · n=195 (stratified) · n=50 · pilots (n=5/10/1). Only same-surface runs are comparable.")


def slide_llama(wb: openpyxl.Workbook, llama: dict | None, llama_sorter: dict | None) -> None:
    ws, row = new_slide(
        wb, "PART B · SORTER", "Llama runs (Langfuse)", 16,
        kicker="llama sorter run landed 2026-08-16 (KANBAN-042): meta-llama/llama-3.3-70b-instruct × sorter_v13, full-509; plus the earlier llama-4-scout EXTRACTION run from Langfuse",
    )
    if llama_sorter is not None:
        sc = llama_sorter["scores"]["sorter"]
        ci = sc.get("subtype_accuracy_ci", {})
        row = section_row(ws, row, "Llama sorter run — meta-llama/llama-3.3-70b-instruct × sorter_v13 (full-509, KANBAN-042)")
        row = table(ws, row, 2,
                    ["Metric", "Llama-3.3-70B (n=509)", "Qwen champion (n=509)", "DeepSeek V4 Pro (n=509)"],
                    [
                        ["Subtype accuracy", _num(sc.get("subtype_accuracy")), _num(0.9430), _num(0.9528)],
                        ["Exact match (doc_type + subtype)", _num(sc.get("exact_match")), _num(0.9961), _num(0.9961)],
                        ["Subtype equiv", _num(sc.get("subtype_accuracy_equiv")), _num(0.9470), _num(0.9548)],
                        ["Mean confidence", _num(sc.get("confidence")), _num(0.9583), _num(0.9555)],
                        ["Failed rows", sc.get("failure_insights", {}).get("n_failed"), 29, 24],
                    ],
                    widths=[2, 26, 20, 26, 26, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 2],
                    highlight_cols={2, 3, 4})
        row = note(ws, row, 2, f"95% CI [{_num(ci.get('lo'))}, {_num(ci.get('hi'))}] — 62 fails: 47 family_confusion / 11 equivalent_family / 3 function_over_form / 1 other_fallback. 6.66M tokens, est. cost None (OpenRouter-billed, no local price).")
        row += 1
    row = section_row(ws, row, "Earlier llama run — llama-4-scout × contracts_specialist_v31 EXTRACTION (Langfuse llm-dojo)")
    row = kv_table(ws, row, 2, [
        ("Session", llama.get("session_id") if llama else "llama-4-scout_contracts_specialist_v31_extraction_langfuse"),
        ("Task / prompt", f"{llama.get('task') if llama else 'contract_entity_extraction'} · {llama.get('prompt') if llama else 'contracts_specialist_v31'}"),
        ("Model", llama.get("model") if llama else "llama-4-scout (OpenRouter)"),
        ("Traces / scored", f"{llama.get('n_traces') if llama else 509} doc traces · only {llama.get('n_scored') if llama else 20} scored (run truncated)"),
        ("Window", llama.get("window") if llama else "2026-08-16 06:55–07:22 UTC"),
        ("Provenance", "Langfuse llm-dojo traces + scores (sourced via langfuse-cli; NOT in reports/experiment_log.jsonl)"),
    ])
    row = section_row(ws, row, "llama-4-scout extraction headline scores (n=20 scored docs)")
    d = llama or {}
    sc4 = d.get("scores", {})
    row = table(ws, row, 2,
                ["Metric", "Llama-4-scout (n=20)", "Qwen 3.7-Flash v31 (n=509)", "Qwen 3.7-Flash v32 (n=509)"],
                [
                    ["Overall extraction score", _num(sc4.get("overall_extraction_score", 0.6627)), _num(0.8737), _num(0.8807)],
                    ["Field presence", _num(sc4.get("field_presence", 0.8935)), _num(0.9655), _num(0.9701)],
                    ["Schema valid", _num(sc4.get("schema_valid", 1.0)), _num(1.0), _num(1.0)],
                    ["Verified precision", _num(sc4.get("overall_verified_precision", 0.9589)), _num(0.9799), _num(0.9799)],
                    ["Category presence", _num(sc4.get("category_presence", 0.4782)), _num(0.8474), _num(0.8555)],
                ],
                widths=[2, 26, 20, 26, 26, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 2],
                highlight_cols={2, 3, 4})
    row = section_row(ws, row, "Usage (llama-4-scout run)")
    row = kv_table(ws, row, 2, [
        ("Tokens (619 generations incl. chunk windows)", f"{d.get('tokens', 10337518):,} total"),
        ("Cost recorded", "0 (no Langfuse price for the model — OpenRouter-billed)"),
    ])
    note(ws, row, 2, "⚠ llama-4-scout extraction numbers NOT comparable to the qwen rows: n=20 vs 509, run truncated mid-scoring. Treat as signal only.")
    note(ws, row, 2, "Sweep status (sorter_v13, full-509): qwen3.7-flash 0.9430 · deepseek-v4-pro 0.9528 · deepseek-v4-flash 0.9253 · gpt-5-nano 0.8978 · llama-3.3-70b 0.8782 · gpt-4.1-nano pending. The llama-3.3-70b sorter run traces to Langfuse llm-dojo (KANBAN-042, research-funding key).")


def slide_codebook_subtypes(wb: openpyxl.Workbook) -> None:
    ws, row = new_slide(
        wb, "CODEBOOK · SORTER", "The 25 CUAD contract subtypes", 17,
        kicker="agents/sorter_agent.py CONTRACT_SUBTYPES — the option list IS the schema enum (pinned by test)",
    )
    half = (len(CONTRACT_SUBTYPES) + 1) // 2
    left, right = CONTRACT_SUBTYPES[:half], CONTRACT_SUBTYPES[half:]
    hdr = ["Key", "Label"]
    r1 = table(ws, row, 2, hdr, [[s["key"], s["label"]] for s in left], widths=[2, 18, 30, 10, 10, 12, 11, 11, 11, 11, 11, 11, 11, 11, 11, 2])
    r2 = table(ws, row, 6, hdr, [[s["key"], s["label"]] for s in right], widths=[2, 18, 30, 10, 10, 12, 11, 11, 11, 11, 11, 11, 11, 11, 11, 2])
    row = max(r1, r2) + 1
    row = section_row(ws, row, "Family definitions (operative descriptions)")
    dl, dr = left, right
    rows_dl = [[s["key"], s["description"]] for s in dl]
    rows_dr = [[s["key"], s["description"]] for s in dr]
    hdr2 = ["Key", "Definition"]
    r1 = table(ws, row, 2, hdr2, rows_dl, widths=[2, 14, 44, 10, 10, 12, 11, 11, 11, 11, 11, 11, 11, 11, 11, 2])
    r2 = table(ws, row, 6, hdr2, rows_dr, widths=[2, 14, 44, 10, 10, 12, 11, 11, 11, 11, 11, 11, 11, 11, 11, 2])
    row = max(r1, r2)
    note(ws, row, 2, "Hybrids (\"distribution and development agreement\") can plausibly be either — that is what subtype_accuracy_equiv and the confusion matrix are for (strict stays the discriminating signal).")


def slide_codebook_sorter_rules(wb: openpyxl.Workbook, fs: dict[str, Any]) -> None:
    ws, row = new_slide(
        wb, "CODEBOOK · SORTER", "Equivalences, failure modes & scoring rules", 18,
        kicker="how strict vs equivalent accuracy is computed, and how failures are classified",
    )
    row = section_row(ws, row, "Subtype equivalence families (SUBTYPE_EQUIVALENCES)")
    eqs = [" ↔ ".join(sorted(cls)) for cls in SUBTYPE_EQUIVALENCES]
    row = table(ws, row, 2, ["Family pair (interchangeable for equiv accuracy)"], [[e] for e in eqs], widths=[2, 60, 40, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 2])
    row += 1
    row = section_row(ws, row, "Failure-mode taxonomy (failure_insights)")
    row = table(ws, row, 2,
                ["Mode", "Definition"],
                [
                    ["function_over_form", "doc_type miss — document classified under the wrong primary class"],
                    ["other_fallback", "model fell back to 'other' instead of the true family"],
                    ["equivalent_family", "wrong subtype, same equivalence family — recovered in equiv accuracy"],
                    ["family_confusion", "wrong family tier, not recoverable — the strict-accuracy loss"],
                ],
                widths=[2, 24, 76, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 2])
    row += 1
    row = section_row(ws, row, "Scoring rules")
    row = table(ws, row, 2,
                ["Rule", "Setting"],
                [
                    ["exact_match", "doc_type AND subtype both correct (strict, row-level)"],
                    ["subtype_accuracy", "strict — exact CUAD-folder key; the discriminating signal"],
                    ["subtype_accuracy_equiv", "strict OR equivalent_subtypes() — honors the 4 family classes"],
                    ["Confidence gate", f"confidence ≥ {fs.get('confidence', {}).get('high', 0.95)} auto-continues; < {fs.get('confidence', {}).get('low', 0.7)} retry → human review"],
                    ["Option list == schema enum", "SorterAgent's CONTRACT_SUBTYPES list and the JSON-schema enum must be identical (enforced by test)"],
                ],
                widths=[2, 26, 74, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 2])
    note(ws, row, 2, "Reasoning effort: sorter defaults to 'medium' (25 near-synonymous families need deliberation — verified +4.6pp strict on the 200-doc sample).")


def slide_docclass_and_sources(wb: openpyxl.Workbook, r: dict, records: dict[str, dict]) -> None:
    ws, row = new_slide(
        wb, "PART C · DOCCLASS v5", "Hierarchical doc-class diag + sources", 19,
        kicker=f"{_model(r)} × sorter_docclass_v5 — 30-doc diagnostic over the MERGED surface (CUAD + MAUD + S-1) — diagnostic, NOT comparable to 509-doc runs",
    )
    sc = r["scores"]
    mc_ = r["scores"]["sorter"]["failure_insights"]["mode_counts"]
    fails_str = " · ".join(f"{k} {v}" for k, v in mc_.items())
    row = table(ws, row, 2,
                ["Metric", "Value"],
                [
                    ["doc_type accuracy", f"{_num(sc.get('doc_type_accuracy'))}  [{_num(sc.get('doc_type_accuracy_ci', {}).get('lo'))}, {_num(sc.get('doc_type_accuracy_ci', {}).get('hi'))}]  (n=30)"],
                    ["subclass accuracy", f"{_num(sc.get('subclass_accuracy'))}  [{_num(sc.get('subclass_accuracy_ci', {}).get('lo'))}, {_num(sc.get('subclass_accuracy_ci', {}).get('hi'))}]  (n={sc.get('subclass_accuracy_ci', {}).get('n')} subclass-GT rows)"],
                    ["exact match (doc_type + subclass)", f"{_num(sc.get('exact_match'))}  [{_num(sc.get('exact_match_ci', {}).get('lo'))}, {_num(sc.get('exact_match_ci', {}).get('hi'))}]"],
                    ["mean confidence", _num(sc.get("confidence"))],
                    ["Failures", f"{sc.get('sorter', {}).get('failure_insights', {}).get('n_failed')}  ({fails_str})"],
                    ["Tokens / est. cost", f"{r.get('tokens', {}).get('total', {}).get('total_tokens', 0):,} · ${r.get('tokens', {}).get('total', {}).get('cost_estimated_usd', 0):.4f}"],
                ],
                widths=[2, 30, 70, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 2])
    row += 1
    row = section_row(ws, row, "Per-class & per-subclass accuracy")
    pc = sc.get("per_class_accuracy", {})
    ps = sc.get("per_subclass_accuracy", {})
    sup = sc.get("per_subclass_support", {})
    rows_pc = [[k, _num(v)] for k, v in pc.items()]
    rows_ps = [[k, _num(v), sup.get(k)] for k, v in ps.items()]
    r1 = table(ws, row, 2, ["Class", "Acc"], rows_pc, widths=[2, 28, 10, 10, 10, 12, 11, 11, 11, 11, 11, 11, 11, 11, 11, 2])
    r2 = table(ws, row, 6, ["Subclass", "Acc", "n"], rows_ps, widths=[2, 28, 10, 10, 10, 12, 11, 11, 11, 11, 11, 11, 11, 11, 11, 2])
    row = max(r1, r2) + 1
    row = section_row(ws, row, "Sources & navigation")
    row = table(ws, row, 2,
                ["Source", "Location"],
                [
                    ["Experiment log (append-only JSONL)", "reports/experiment_log.jsonl"],
                    ["Rendered log (derived — never hand-edit)", "reports/experiment_log.md (python scripts/reporting/render_experiment_log.py)"],
                    ["Extraction run record", f"{RUN_NAMES['extraction']} · git {records[RUN_NAMES['extraction']].get('git', {}).get('commit')}"],
                    ["Sorter subtype run record", f"{RUN_NAMES['sorter']} · git {records[RUN_NAMES['sorter']].get('git', {}).get('commit')}"],
                    ["Docclass run record", f"{RUN_NAMES['docclass']} · git {records[RUN_NAMES['docclass']].get('git', {}).get('commit')}"],
                    ["Prompts", "src/prompts.py (contracts_specialist_v32, SORTER_PROMPT_V14, sorter_docclass_v5)"],
                    ["Taxonomy / thresholds", "config/taxonomy.yaml"],
                    ["Scoring model (formula-level reference)", "docs/SCORING.md"],
                    ["Sorter subtypes / equivalences", "agents/sorter_agent.py"],
                    ["Memos", "docs/memos/contracts_specialist_v32.md · docs/memos/sorter_v14.md · docs/memos/docclass_v3_merged_benchmark.md"],
                ],
                widths=[2, 40, 60, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 2])
    note(ws, row, 2, "Docclass context: sorter_docclass_v3 = completed prompt (doc_type 0.9926 / subclass 0.5808 / exact 0.8905 on the merged 676); v5 is the diagnostic iteration — small n, wide CIs, read as signal only.")


# ---------------------------------------------------------------------------
# Builder
# ---------------------------------------------------------------------------


def load_llama_run(json_path: str) -> dict | None:
    """Load the Langfuse-sourced llama run record (None when absent)."""
    if not os.path.exists(json_path):
        return None
    with open(json_path, encoding="utf-8") as fh:
        return json.load(fh)


def build_deck(log_path: str, taxonomy_path: str, out_path: str,
               llama_json: str = "reports/sheets/llama4scout_v31_extraction_langfuse.json") -> str:
    """Build the deck workbook; returns the output path."""
    records = load_records(log_path)
    records_list = load_records_list(log_path)
    missing = [n for n in RUN_NAMES.values() if n not in records]
    if missing:
        raise SystemExit(f"Missing experiment-log records: {missing} — is {log_path} current?")
    extraction = records[RUN_NAMES["extraction"]]
    sorter = records[RUN_NAMES["sorter"]]
    docclass = records[RUN_NAMES["docclass"]]
    fields = load_contract_fields(taxonomy_path)
    fs = load_field_scoring(taxonomy_path)
    llama = load_llama_run(llama_json)
    llama_sorter = records.get("meta-llama-llama-3.3-70b-instruct_sorter_v13_subtype_langfuse")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    slide_cover(wb)
    slide_contents(wb)
    slide_extract_metadata(wb, extraction)
    slide_extract_headlines(wb, extraction)
    slide_extract_per_field(wb, extraction)
    slide_extract_errors(wb, extraction)
    slide_extract_entity_lists(wb, extraction)
    slide_extract_diagnostics(wb, extraction)
    slide_codebook_fields(wb, fields)
    slide_codebook_scoring(wb, fs)
    slide_sorter_headlines(wb, sorter)
    slide_sorter_failures(wb, sorter)
    slide_sorter_per_subtype(wb, sorter)
    slide_qwen_lineage_summary(wb, records_list)
    slide_qwen_lineage_runs(wb, records_list)
    slide_llama(wb, llama, llama_sorter)
    slide_codebook_subtypes(wb)
    slide_codebook_sorter_rules(wb, fs)
    slide_docclass_and_sources(wb, docclass, records)

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)
    return out_path


def main_with_args(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--log", default=DEFAULT_LOG, help="experiment log JSONL path")
    parser.add_argument("--taxonomy", default=DEFAULT_TAXONOMY, help="taxonomy YAML path")
    parser.add_argument("--outdir", default=DEFAULT_OUTDIR, help="output directory")
    parser.add_argument("--outfile", default=DEFAULT_DECK, help="output workbook filename")
    parser.add_argument("--llama-json",
                        default="reports/sheets/llama4scout_v31_extraction_langfuse.json",
                        help="Langfuse-sourced llama run record (optional)")
    args = parser.parse_args(argv)
    out = os.path.join(args.outdir, args.outfile)
    built = build_deck(args.log, args.taxonomy, out, args.llama_json)
    wb = openpyxl.load_workbook(built, read_only=True)
    print(f"Wrote {built} — {len(wb.sheetnames)} slides: {', '.join(wb.sheetnames)}")
    return 0


def main() -> None:
    raise SystemExit(main_with_args(sys.argv[1:]))


if __name__ == "__main__":
    main()
