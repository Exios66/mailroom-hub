"""Unit tests for the model-sweep exporter (scripts/reporting/export_sweep_results.py).

Network-free: builds tiny synthetic experiment-log records and verifies the
filter (champion-prompt runs only, chronological), the Notes column, and the
reference-compatible workbook structure (114 reference columns + Notes,
Eval Results + Codebook sheets).
"""

from __future__ import annotations

import openpyxl


def _record(name: str, ts: str, prompt: str = "sorter_v13", n: int = 509,
            model: str = "qwen/qwen3.7-flash", subtype: float = 0.9,
            task: str = "subtype_classification") -> dict:
    return {
        "task": task,
        "experiment_name": name,
        "model": model,
        "timestamp": ts,
        "n_rows": n,
        "n_ok": n,
        "prompt_versions": {"sorter": prompt},
        "parameters": {"temperature": 0.1, "reasoning_effort": "medium",
                       "max_tokens": 4096, "max_input_chars": 100000,
                       "max_concurrency": 8},
        "data_source": {"project": "llm-mailroom/mailroom-cuad-contracts-full",
                        "n_samples": n, "sample_requested": 0,
                        "stratified": 0, "seed": 42},
        "tokens": {"sorter": {"prompt_tokens": 1000, "completion_tokens": 100,
                              "total_tokens": 1100, "cost_usd": 0,
                              "cost_total_usd": 0, "cost_estimated_usd": 0.05,
                              "rows_with_usage": n}},
        "scores": {"sorter": {
            "exact_match": 0.99, "exact_match_ci": {"lo": 0.98, "hi": 1.0, "half": 0.01},
            "subtype_accuracy": subtype,
            "subtype_accuracy_ci": {"lo": subtype - 0.02, "hi": subtype + 0.02, "half": 0.02},
            "subtype_accuracy_equiv": subtype, "confidence": 0.95,
            "failure_insights": {"n_failed": 5, "mode_counts": {
                "equivalent_family": 1, "family_confusion": 3,
                "function_over_form": 1, "other_fallback": 0}},
            "per_subtype": {},
        }},
    }


def _sweep_rows(records: list[dict]) -> list[dict]:
    from scripts.reporting.export_sweep_results import sweep_records  # type: ignore

    return sweep_records(records, "sorter_v13")


def test_sweep_filters_champion_prompt_and_sorts():
    rows = _sweep_rows([
        _record("b_sorter_v13", "2026-08-16T05:00:00+00:00"),
        _record("a_sorter_v12", "2026-08-16T04:00:00+00:00", prompt="sorter_v12"),
        _record("a_sorter_v13", "2026-08-16T03:00:00+00:00"),
    ])
    assert [r["experiment_name"] for r in rows] == ["a_sorter_v13", "b_sorter_v13"]


def test_sweep_excludes_other_tasks():
    rows = _sweep_rows([
        _record("extraction_run", "2026-08-16T03:00:00+00:00", task="contract_entity_extraction"),
        _record("ok_sorter_v13", "2026-08-16T03:00:00+00:00"),
    ])
    assert [r["experiment_name"] for r in rows] == ["ok_sorter_v13"]


def test_run_note_generic_fallback():
    from scripts.reporting.export_sweep_results import run_note  # type: ignore

    assert run_note(_record("x_smoke", "2026-08-16T03:00:00+00:00", n=1)) == (
        "1-doc smoke run (pre-launch gate on the default key)")
    degraded = _record("x_degraded", "2026-08-16T03:00:00+00:00", n=509)
    degraded["n_ok"] = 416
    assert "DEGRADED" in run_note(degraded)
    assert run_note(_record("x_full", "2026-08-16T03:00:00+00:00")) == (
        "Full-corpus run on the champion prompt")


def test_sweep_workbook_structure(tmp_path):
    from scripts.reporting.export_sweep_results import (  # type: ignore
        build_sweep_workbook,
    )

    records = [
        _record("qwen_champ", "2026-08-16T03:00:00+00:00"),
        _record("nano_bench", "2026-08-16T04:00:00+00:00", model="openai/gpt-5-nano"),
    ]
    log = tmp_path / "log.jsonl"
    log.write_text("\n".join(__import__("json").dumps(r) for r in records) + "\n")
    out = tmp_path / "Sorter_Model_Sweep_Results.xlsx"
    n = build_sweep_workbook(str(log), "sorter_v13", str(out))
    assert n == 2

    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["Eval Results", "Codebook"]
    ws = wb["Eval Results"]
    assert ws.max_column == 115  # 114 reference columns + Notes
    headers = [c.value for c in ws[1]]
    assert headers[:5] == ["DATE", "Experiment Name", "SAMPLE (n)", "MODEL", "Prompt Version"]
    assert headers[-1] == "Notes"
    assert ws.max_row == 3
    assert ws.freeze_panes == "F2"
    assert ws[1][0].fill.fgColor.rgb == "001F4E79"
    cb = wb["Codebook"]
    assert cb.max_row == 116  # 115 variables + header
    assert [c.value for c in cb[1]] == ["Variable", "Description", "Type", "Source", "Example / Values"]
