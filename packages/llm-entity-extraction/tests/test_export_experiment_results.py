"""Unit tests for the experiment-results exporter (scripts/reporting/export_experiment_results.py).

Network-free: builds tiny synthetic experiment-log records and verifies the
workbook/codebook structure (columns, formatting, sheet layout) matches the
reference format contract.
"""

from __future__ import annotations

import csv
import json

import openpyxl


def _sorter_record(name="qwen3.7-flash_sorter_v12_subtype_langfuse") -> dict:
    return {
        "type": "experiment",
        "task": "subtype_classification",
        "experiment_name": name,
        "model": "qwen/qwen3.7-flash",
        "timestamp": "2026-08-15T22:19:00+00:00",
        "n_rows": 509,
        "n_ok": 509,
        "prompt_versions": {"sorter": "sorter_v12"},
        "parameters": {"temperature": 0.1, "reasoning_effort": "medium",
                       "max_tokens": 4096, "max_input_chars": 100000,
                       "max_concurrency": 8},
        "data_source": {"project": "llm-mailroom/mailroom-cuad-contracts-full",
                        "n_samples": 509, "sample_requested": 0,
                        "stratified": 0, "seed": 42},
        "tokens": {"sorter": {"prompt_tokens": 1000, "completion_tokens": 100,
                              "total_tokens": 1100, "cost_usd": 0,
                              "cost_total_usd": 0, "cost_estimated_usd": 0.05,
                              "rows_with_usage": 509},
                   "total": {"prompt_tokens": 1000, "completion_tokens": 100,
                             "total_tokens": 1100, "cost_usd": 0,
                             "cost_total_usd": 0, "cost_estimated_usd": 0.05,
                             "rows_with_usage": 509}},
        "scores": {
            "sorter": {
                "exact_match": 0.9961,
                "exact_match_ci": {"lo": 0.9902, "hi": 1.0, "half": 0.0049},
                "subtype_accuracy": 0.9234,
                "subtype_accuracy_ci": {"lo": 0.8978, "hi": 0.945, "half": 0.0236},
                "subtype_accuracy_equiv": 0.9312,
                "confidence": 0.9556,
                "failure_insights": {
                    "n_failed": 39,
                    "mode_counts": {"equivalent_family": 4, "family_confusion": 29,
                                    "function_over_form": 2, "other_fallback": 4},
                },
                "per_subtype": {
                    "affiliate": {"accuracy": 0.9, "accuracy_equiv": 1.0,
                                  "correct": 9, "equiv": 10, "total": 10},
                },
            },
        },
    }


def _extraction_record(name="qwen3.7-flash_contracts_specialist_v32_extraction") -> dict:
    return {
        "type": "experiment",
        "task": "contract_entity_extraction",
        "experiment_name": name,
        "model": "qwen/qwen3.7-flash",
        "timestamp": "2026-08-16T02:52:00+00:00",
        "n_rows": 509,
        "n_ok": 509,
        "prompt_version": "contracts_specialist_v32",
        "parameters": {"temperature": 0.1, "reasoning_effort": "none",
                       "max_tokens": 16384, "max_input_chars": 150000,
                       "max_concurrency": 8, "judge": False, "chunked": True,
                       "chunk_chars": 90000, "chunk_overlap": 8000},
        "data_source": {"project": "llm-mailroom/mailroom-cuad-contracts-full",
                        "n_samples": 509, "seed": 42},
        "tokens": {"prompt_tokens": 6000000, "completion_tokens": 400000,
                   "total_tokens": 6400000, "cost_usd": 0,
                   "cost_total_usd": 0, "cost_estimated_usd": 0.25,
                   "rows_with_usage": 509},
        "scores": {
            "overall_extraction_score": 0.887,
            "overall_extraction_score_ci": {"lo": 0.8689, "hi": 0.8913, "half": 0.0112},
            "field_presence": 0.9701,
            "schema_valid": 1.0,
            "category_presence": 0.8555,
            "overall_verified_precision": 0.9819,
            "hallucination_rate": {"document_name": 0.0, "effective_date": 0.0546,
                                   "governing_law": 0.0, "key_obligations": 0.0021,
                                   "parties": 0.0348, "renewal_terms": 0.0,
                                   "term_length": 0.0212, "termination_clauses": 0.0045},
            "verified_precision": {"document_name": 1.0, "effective_date": 0.9454,
                                   "governing_law": 1.0, "key_obligations": 0.9979,
                                   "parties": 0.9652, "renewal_terms": 1.0,
                                   "term_length": 0.9788, "termination_clauses": 0.9955},
            "entity_list_f1": {"key_obligations": 0.7881, "parties": 0.9038,
                               "termination_clauses": 0.8933},
            "per_field": {"document_name": 0.9874, "effective_date": 0.8877,
                          "governing_law": 0.929, "key_obligations": 0.781,
                          "parties": 0.9167, "renewal_terms": 0.8124,
                          "term_length": 0.8026, "termination_clauses": 0.9012},
            "diagnostics": {
                "n_fields_scored": 3121,
                "field_exact_rate": 0.7632,
                "field_partial_rate": 0.1839,
                "field_miss_rate": 0.0529,
                "list_precision": 0.4362,
                "list_recall": 0.7881,
                "list_f1": 0.5122,
                "list_micro_precision": 0.3263,
                "list_micro_recall": 0.7875,
                "list_micro_f1": 0.4614,
                "list_micro_matched": 3164,
                "list_micro_n_expected": 4018,
                "list_micro_n_predicted": 9696,
                "entity_list_precision": {"key_obligations": 0.4362, "parties": 0.3874,
                                          "termination_clauses": 0.4583},
                "entity_list_recall": {"key_obligations": 0.7881, "parties": 0.9038,
                                       "termination_clauses": 0.8933},
                "entity_list_raw_f1": {"key_obligations": 0.5122, "parties": 0.5299,
                                       "termination_clauses": 0.5622},
                "date_mae_days": 34.184, "date_median_ae_days": 0.0, "date_r2": 0.9824,
                "date_n_pairs": 413,
                "date_mae_per_field": {"effective_date": 33.3811, "term_length": 365.0},
                "date_r2_per_field": {"effective_date": 0.9824, "term_length": None},
                "duration_mae_days": 423.9459, "duration_median_ae_days": 0.0,
                "duration_r2": 0.7315, "duration_n_pairs": 148,
                "duration_mae_per_field": {"renewal_terms": 431.042, "term_length": 394.8276},
                "duration_r2_per_field": {"renewal_terms": -0.2386, "term_length": 0.8983},
                "money_n_pairs": 0,
                "span_count_n_docs": 1129, "span_count_mae": 5.3481,
                "span_count_signed_mean": 5.0292,
                "span_count_mae_per_field": {"parties": 2.0461, "key_obligations": 10.3916,
                                             "termination_clauses": 1.7978},
                "span_count_signed_mean_per_field": {"parties": 2.0461,
                                                     "key_obligations": 9.6128,
                                                     "termination_clauses": 1.7528},
                "field_presence_per_field": {"document_name": 0.998, "parties": 1.0,
                                             "effective_date": 0.886, "term_length": 0.83,
                                             "termination_clauses": 0.87, "governing_law": 0.86,
                                             "key_obligations": 0.986, "contract_value": 0.4,
                                             "renewal_terms": 0.374},
                "error_decomposition": {"document_name": {"exact_rate": 0.958, "partial_rate": 0.04, "miss_rate": 0.002},
                                        "effective_date": {"exact_rate": 0.8202, "partial_rate": 0.0992, "miss_rate": 0.0806},
                                        "governing_law": {"exact_rate": 0.8555, "partial_rate": 0.1282, "miss_rate": 0.0163},
                                        "key_obligations": {"exact_rate": 0.3341, "partial_rate": 0.6438, "miss_rate": 0.0221},
                                        "parties": {"exact_rate": 0.9038, "partial_rate": 0.0, "miss_rate": 0.0962},
                                        "renewal_terms": {"exact_rate": 0.6149, "partial_rate": 0.2931, "miss_rate": 0.092},
                                        "term_length": {"exact_rate": 0.6691, "partial_rate": 0.2691, "miss_rate": 0.0617},
                                        "termination_clauses": {"exact_rate": 0.8933, "partial_rate": 0.0, "miss_rate": 0.1067}},
            },
        },
    }


def _write_log(tmp_path, records) -> str:
    p = tmp_path / "experiment_log.jsonl"
    with open(p, "w") as fh:
        for r in records:
            fh.write(json.dumps(r) + "\n")
    return str(p)


def test_sorter_workbook_structure(tmp_path):
    from scripts.reporting.export_experiment_results import (
        main_with_args, sorter_columns, sorter_records,
    )

    log = _write_log(tmp_path, [_sorter_record()])
    out = tmp_path / "out"
    rc = main_with_args(["--task", "sorter", "--outdir", str(out), "--log", log])
    assert rc == 0

    wb = openpyxl.load_workbook(out / "Sorter_Experiment_Results.xlsx")
    assert wb.sheetnames == ["Eval Results", "Codebook"]
    ws = wb["Eval Results"]
    assert ws.max_column == 114
    assert ws.max_row == 2
    assert ws.freeze_panes == "F2"
    assert ws.auto_filter.ref == "A1:DJ2"
    # header + formatting
    assert ws.cell(row=1, column=2).value == "Experiment Name"
    assert ws.cell(row=1, column=2).font.bold
    assert ws.cell(row=2, column=1).number_format == "mm/dd/yyyy"
    assert ws.cell(row=2, column=8).number_format == "0.00%"
    # data row values
    assert ws.cell(row=2, column=2).value == "qwen3.7-flash_sorter_v12_subtype_langfuse"
    assert ws.cell(row=2, column=5).value == "v12"  # derived from sorter_v12
    assert ws.cell(row=2, column=4).value == "Qwen 3.7-Flash"
    assert abs(ws.cell(row=2, column=8).value - 0.9234) < 1e-9
    assert ws.cell(row=2, column=24).value == 0.9  # Accuracy: affiliate
    # codebook sheet
    cb = wb["Codebook"]
    assert cb.max_column == 5
    assert cb.cell(row=1, column=1).value == "Variable"

    # codebook CSV: one row per column
    with open(out / "Sorter_Experiment_Codebook.csv") as fh:
        rows = list(csv.reader(fh))
    assert len(rows) == 115
    assert rows[0] == ["Variable", "Description", "Type", "Source", "Example / Values"]
    assert all(len(r) == 5 for r in rows)
    hdrs = [ws.cell(row=1, column=c).value for c in range(1, 115)]
    assert {r[0] for r in rows[1:]} == set(hdrs)


def test_extraction_workbook_structure(tmp_path):
    from scripts.reporting.export_experiment_results import main_with_args

    log = _write_log(tmp_path, [_extraction_record()])
    out = tmp_path / "out"
    rc = main_with_args(["--task", "extraction", "--outdir", str(out), "--log", log])
    assert rc == 0

    wb = openpyxl.load_workbook(out / "Entity_Extraction_Results.xlsx")
    assert wb.sheetnames == ["Eval Results"]
    ws = wb["Eval Results"]
    assert ws.max_column == 141
    assert ws.max_row == 2
    assert ws.freeze_panes == "F2"
    # values
    assert ws.cell(row=2, column=5).value == "v32"
    assert abs(ws.cell(row=2, column=7).value - 0.887) < 1e-9
    assert ws.cell(row=2, column=135).value == 1  # chunked
    assert ws.cell(row=2, column=138).value == "False"  # judge as string (reference format)
    # diagnostics columns
    assert abs(ws.cell(row=2, column=42).value - 0.7632) < 1e-9  # field exact rate
    assert abs(ws.cell(row=2, column=55).value - 0.4362) < 1e-9  # list precision: key_obligations
    assert ws.cell(row=2, column=67).value == 413  # date n pairs
    assert abs(ws.cell(row=2, column=81).value - 5.3481) < 1e-9  # span count MAE
    assert abs(ws.cell(row=2, column=95).value - 0.374) < 1e-9  # field presence: renewal_terms
    assert abs(ws.cell(row=2, column=96).value - 0.83) < 1e-9  # field presence: term_length
    assert abs(ws.cell(row=2, column=98).value - 0.4) < 1e-9  # field presence: contract_value
    assert abs(ws.cell(row=2, column=122).value - 0.1067) < 1e-9  # err miss: termination_clauses

    # codebook CSV
    with open(out / "Entity_Extraction_Codebook.csv") as fh:
        rows = list(csv.reader(fh))
    assert len(rows) == 142
    hdrs = [ws.cell(row=1, column=c).value for c in range(1, 142)]
    assert {r[0] for r in rows[1:]} == set(hdrs)
