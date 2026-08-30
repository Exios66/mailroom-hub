"""Integration tests for the llm-dojo-scoring dependency (KANBAN-044).

The scoring/error-analysis/export definitions now live in the
``llm_dojo_scoring`` package; the local ``src/`` modules are re-export shims.
These tests pin the integration contract:

- the package ``Settings`` are wired from ``config/taxonomy.yaml``
  (embedding_enabled, cost_models dict-form conversion, thresholds);
- the local shims keep their one-argument / keyword contracts
  (``get_field_types``, ``extraction_diagnostics(master=...)``,
  ``build_scorers``/``cost``/object-list ``per_class_stats``);
- the exported workbook column specs are byte-identical to the package's
  (and to the committed reference-format artifacts);
- the ``dojo-analyze`` CLI works on this repo's workbook artifacts.

All tests are network-free.
"""

from __future__ import annotations

import json
import subprocess
import sys

import openpyxl

from llm_dojo_scoring.export import (
    extraction_columns as _package_extraction_columns,
)
from llm_dojo_scoring.export import sorter_columns as _package_sorter_columns
from llm_dojo_scoring.field_scoring import get_settings

from src.field_scoring import (
    get_ambiguous_band,
    get_bipartite_match_threshold,
    get_containment_fields,
    get_field_types,
    get_partial_gt_fields,
    is_entity_list,
    score_extraction,
)
from src.metrics import extraction_diagnostics
from src.scorers import build_scorers, cost, macro_accuracy, per_class_stats

REPO_ROOT = __import__("pathlib").Path(__file__).resolve().parents[1]


# ---------------------------------------------------------------------------
# Settings wiring (config/taxonomy.yaml -> package Settings)
# ---------------------------------------------------------------------------


def test_taxonomy_settings_wired_into_package():
    settings = get_settings()
    assert settings.field_scoring.embedding_enabled is True  # taxonomy: true (package default False)
    assert settings.field_scoring.ambiguous_band == (0.5, 0.85)
    assert settings.field_scoring.bipartite_match_threshold == 0.6
    assert settings.field_scoring.verification_enabled is True
    assert settings.field_scoring.verification_token_coverage == 0.7
    assert settings.field_scoring.partial_gt_fields == {"parties", "key_obligations", "termination_clauses"}
    assert settings.field_scoring.containment_fields == {"governing_law", "term_length", "renewal_terms"}


def test_cost_models_dict_form_converted():
    # taxonomy.yaml writes {model: {input_per_million, output_per_million}};
    # the package wants [in, out] lists. The converted table must match the
    # YAML prices exactly.
    settings = get_settings()
    assert settings.cost_models["qwen/qwen3.7-flash"] == (0.03, 0.13)
    assert settings.cost_models["deepseek/deepseek-v4-flash"] == (0.05, 0.25)
    assert settings.cost_models["deepseek/deepseek-v4-pro"] == (0.435, 0.87)
    from src.cost_models import estimate_cost

    assert estimate_cost(1_000_000, 0, "qwen/qwen3.7-flash") == 0.03


def test_config_accessors_match_package_settings():
    assert get_ambiguous_band() == (0.5, 0.85)
    assert get_bipartite_match_threshold() == 0.6
    assert get_partial_gt_fields() == {"parties", "key_obligations", "termination_clauses"}
    assert get_containment_fields() == {"governing_law", "term_length", "renewal_terms"}


# ---------------------------------------------------------------------------
# Shim contracts
# ---------------------------------------------------------------------------


def test_get_field_types_one_arg_resolves_repo_taxonomy():
    field_types = get_field_types("contract")
    # The taxonomy writes compound entity-list types ("entity_list:free_text");
    # the contract is that the value resolves to entity-list scoring (the base
    # type of the compound), exactly as the old local field_types() behaved.
    assert is_entity_list(field_types["key_obligations"])
    assert field_types["key_obligations"] == "entity_list:free_text"
    assert field_types["effective_date"] == "date"
    assert get_field_types("no_such_class") == {}


def test_score_extraction_returns_package_result_shape():
    field_types = get_field_types("contract")
    result = score_extraction(
        "contract", field_types,
        {"document_name": "Acme Corp", "effective_date": "2024-03-03"},
        {"document_name": "Acme Corp", "effective_date": "2024-03-03"},
    )
    assert result.overall_score == 1.0
    assert result.field_scores["document_name"] == 1.0
    assert hasattr(result, "to_dict")  # package superset


def _sorter_record(filename: str = "doc_1") -> dict:
    return {
        "type": "experiment",
        "task": "subtype_classification",
        "experiment_name": "qwen3.7-flash_sorter_v13_subtype_langfuse",
        "git": {"commit": "test", "dirty": False},
        "model": "qwen/qwen3.7-flash",
        "prompt_version": "sorter_v13",
        "parameters": {"reasoning_effort": "medium", "temperature": 0.1, "seed": 42},
        "timestamp": "2026-08-16T00:00:00Z",
        "n_rows": 2,
        "scores": {"sorter": {"subtype_accuracy": 1.0, "exact_match": 1.0,
                              "n_rows": 2}},
        "results": [
            {"filename": filename, "expected": "license", "predicted": "license",
             "reasoning": "ok"},
        ],
    }


def test_diagnostics_keeps_master_keyword():
    field_types = {"effective_date": "date"}
    rows = [{
        "filename": "contract_1",
        "predicted": {"effective_date": "2024-03-03"},
        "expected_fields": {"effective_date": "March 2, 2024"},
        "field_scores": {"effective_date": 0.5},
        "entity_list_scores": {},
    }]
    master = {
        "contract1": {  # master_labels._norm_filename("contract_1")
            "Agreement Date-Answer": "2024-03-01",  # normalized answer preferred
            "Effective Date-Answer": "2024-03-01",
        },
    }
    diag = extraction_diagnostics(rows, field_types, master=master)
    assert diag["date_n_pairs"] == 1
    assert diag["date_mae_days"] == 2.0  # vs raw expected (2024-03-02 -> 1 day)


def test_scorers_keep_local_registry_contract():
    assert build_scorers(None) and len(build_scorers(None)) == 3
    assert build_scorers(["exact_match", "cost"])[0].__name__ == "exact_match"
    assert cost({}) == 0.0
    assert cost({"cost": 1.5}) == 1.5

    class FakeResult:
        def __init__(self, expected, output):
            self.expected = expected
            self.output = output

    results = [FakeResult("contract", "contract"), FakeResult("contract", "corporate_record")]
    stats = per_class_stats(results)
    assert stats["contract"] == {"n": 2, "correct": 1, "accuracy": 0.5}
    assert macro_accuracy(results) == 0.5


def test_shim_parity_imports_resolve():
    # Every public name the local modules historically exposed must resolve
    # through the shims (llm-mailroom imports these via pip install -e .).
    from src.bootstrap import bootstrap_ci, delta_significance, wilson_ci
    from src.cost_models import estimate_cost, estimate_for_record, price_for, tokens_summary
    from src.experiment_log import append_experiment, git_snapshot, mean, tokens_summary as ts2
    from src.field_scoring import (  # noqa: F401
        EntityListScore,
        ExtractionScoreResult,
        audit_list_field,
        audit_scalar_field,
        is_entity_list,
        normalize_text,
        parse_date,
        parse_money,
        score_category_presence,
        score_containment_field,
        score_date_field,
        score_entity_list,
        score_field,
        score_free_text_field,
        score_id_field,
        score_money_field,
        score_name_field,
        verify_list_items,
    )
    from src.metrics import DATE_FIELDS, DURATION_FIELDS, parse_duration_days
    from src.scorers import ERROR_PREFIX, exact_match, failure, normalize_label

    assert ERROR_PREFIX == "ERROR: "
    assert exact_match("contract", "contract") == 1.0
    assert failure("ERROR: boom", "x") == 1.0
    assert normalize_label("  Contract  ") == "contract"
    assert mean([1.0, 2.0, 3.0]) == 2.0
    assert ts2([{"prompt_tokens": 100, "completion_tokens": 50, "cost": 0.01}])["rows_with_usage"] == 1
    assert git_snapshot()["commit"]  # repo-backed
    assert bootstrap_ci([1.0, 0.0, 1.0, 1.0]) is not None


# ---------------------------------------------------------------------------
# Export byte-identity
# ---------------------------------------------------------------------------


def test_export_columns_byte_identical_to_package():
    # The reporting script must re-export the package's column specs — same
    # function object (identity), so every workbook header stays byte-identical
    # to the package's canonical spec. (Column dicts embed per-call lambdas, so
    # deep equality is meaningless — identity + headers are the contract.)
    import scripts.reporting.export_experiment_results as exp

    assert exp.sorter_columns is _package_sorter_columns
    assert exp.extraction_columns is _package_extraction_columns
    assert [c["header"] for c in exp.sorter_columns()] == [c["header"] for c in _package_sorter_columns()]
    assert [c["header"] for c in exp.extraction_columns()] == [c["header"] for c in _package_extraction_columns()]


def test_export_workbook_headers_match_committed_artifacts(tmp_path):
    # The committed KANBAN-040 sweep workbook shares the 114-column sorter
    # spec plus a trailing reference-format Notes column; its header row must
    # equal the package spec byte-for-byte for the shared 114.
    from scripts.reporting.export_experiment_results import main_with_args, sorter_columns

    workbook_path = REPO_ROOT / "reports" / "sheets" / "Sorter_Model_Sweep_Results.xlsx"
    if not workbook_path.exists():
        import pytest

        pytest.skip("committed sweep workbook absent")
    wb = openpyxl.load_workbook(workbook_path, read_only=True)
    ws = wb["Eval Results"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert headers[:114] == [c["header"] for c in sorter_columns()]
    assert headers[114] == "Notes"

    log = tmp_path / "log.jsonl"
    with log.open("w") as fh:
        fh.write(json.dumps(_sorter_record()) + "\n")
    out = tmp_path / "out"
    rc = main_with_args(["--task", "sorter", "--outdir", str(out), "--log", str(log)])
    assert rc == 0
    gen = openpyxl.load_workbook(out / "Sorter_Experiment_Results.xlsx")
    gws = gen["Eval Results"]
    assert gws.max_column == 114
    assert gws.freeze_panes == "F2"
    assert gws.auto_filter.ref


# ---------------------------------------------------------------------------
# dojo-analyze CLI on this repo's artifacts
# ---------------------------------------------------------------------------


def test_dojo_analyze_cli_smoke(tmp_path):
    from scripts.reporting.export_experiment_results import main_with_args

    log = tmp_path / "log.jsonl"
    with log.open("w") as fh:
        fh.write(json.dumps(_sorter_record()) + "\n")
    out = tmp_path / "out"
    main_with_args(["--task", "sorter", "--outdir", str(out), "--log", str(log)])
    report = tmp_path / "report.md"
    proc = subprocess.run(
        [sys.executable, "-m", "llm_dojo_scoring.cli",
         str(out / "Sorter_Experiment_Results.xlsx"), "-o", str(report), "--no-plots"],
        capture_output=True, text=True, timeout=120,
    )
    assert proc.returncode == 0, proc.stderr
    assert report.exists()
    text = report.read_text()
    assert "subtype" in text.lower() or "champion" in text.lower()
