"""Unit tests for the full results deck exporter
(scripts/reporting/export_full_results_deck.py).

Network-free: builds the deck from synthetic experiment-log records (one
extraction run, one sorter sweep run, two LegalBench runs) and verifies the
19-slide structure, the lineage/sweep/LegalBench tables, and the slide
geometry (landscape fit-to-page sheets with banner + footer).
"""

from __future__ import annotations

import json

import openpyxl


def _record(name: str, ts: str, task: str, n: int, model: str = "qwen/qwen3.7-flash",
            scores: dict | None = None, **score_kw) -> dict:
    return {
        "task": task,
        "type": "experiment",
        "experiment_name": name,
        "model": model,
        "timestamp": ts,
        "n_rows": n,
        "n_ok": n,
        "n_error": 0,
        "prompt_version": "probe_v0",
        "prompt_versions": {"sorter": "probe_v0"},
        "parameters": {"temperature": 0.1, "reasoning_effort": "medium"},
        "data_source": {"source": "synthetic", "n_samples": n},
        "tokens": {"prompt_tokens": 1000, "completion_tokens": 100,
                   "total_tokens": 1100, "cost_usd": 0, "cost_total_usd": 0,
                   "cost_estimated_usd": 0.01, "rows_with_usage": n},
        "scores": scores if scores is not None else score_kw,
    }


def _log(tmp_path) -> str:
    records = [
        _record("qwen3.7-flash_contracts_specialist_v2_extraction",
                "2026-08-09T20:39:28", "extraction", 20,
                overall_extraction_score=0.6563,
                field_presence=0.9, schema_valid=1.0, overall_verified_precision=0.95,
                per_field={"effective_date": {"score": 0.9, "n": 20}},
                diagnostics={"date_mae_days": {"mae": 34.2, "r2": 0.98, "n_pairs": 413}}),
        _record("qwen3.7-flash_contracts_specialist_v32_extraction_langfuse_510_full_clean",
                "2026-08-16T02:52:25", "extraction", 509,
                overall_extraction_score=0.8807,
                overall_extraction_score_ci={"lo": 0.8689, "hi": 0.8913, "half": 0.0112},
                field_presence=0.9701, schema_valid=1.0, overall_verified_precision=0.9799,
                per_field={"effective_date": {"score": 0.9, "n": 500}},
                error_decomposition={"effective_date": {"exact": 0.8, "partial": 0.1, "miss": 0.1}},
                entity_list_scores={"parties": {"gt_coverage": 0.9, "precision": 0.9,
                                                "recall": 0.9, "f1": 0.9}},
                diagnostics={"date_mae_days": {"mae": 34.2, "r2": 0.982, "n_pairs": 413},
                             "span_count_drift": {"mae": 5.35, "signed_mean": 5.03}}),
        _record("qwen3.7-flash_sorter_v13_subtype_langfuse",
                "2026-08-16T03:56:38", "subtype_classification", 509,
                scores={"sorter": {
                    "exact_match": 0.9961, "subtype_accuracy": 0.9430,
                    "subtype_accuracy_ci": {"lo": 0.9214, "hi": 0.9627, "half": 0.0207},
                    "subtype_accuracy_equiv": 0.9470, "confidence": 0.95,
                    "failure_insights": {"n_failed": 29, "mode_counts": {
                        "equivalent_family": 1, "family_confusion": 24,
                        "function_over_form": 2, "other_fallback": 2}},
                    "per_subtype": {}}}),
        _record("deepseek-v4-pro_sorter_v13_subtype_langfuse",
                "2026-08-16T10:46:28", "subtype_classification", 509,
                model="deepseek/deepseek-v4-pro",
                scores={"sorter": {
                    "exact_match": 0.9961, "subtype_accuracy": 0.9528,
                    "subtype_accuracy_ci": {"lo": 0.9332, "hi": 0.9705, "half": 0.0187},
                    "subtype_accuracy_equiv": 0.9548, "confidence": 0.96,
                    "failure_insights": {"n_failed": 24, "mode_counts": {
                        "equivalent_family": 1, "family_confusion": 18,
                        "function_over_form": 2, "other_fallback": 3}},
                    "per_subtype": {}}}),
        _record("qwen3.7-flash_legalbench_task_v2_test", "2026-08-16T02:18:15",
                "task_classification", 94,
                exact_match=0.8830, exact_match_ci={"lo": 0.819, "hi": 0.947, "half": 0.064},
                per_class_accuracy={"no": 0.9057, "yes": 0.8537}),
        _record("qwen3.7-flash_legalbench_task_v3_audit_rights", "2026-08-16T04:10:33",
                "task_classification", 6,
                exact_match=1.0, exact_match_ci={"lo": 1.0, "hi": 1.0, "half": 0.0},
                per_class_accuracy={"no": 1.0, "yes": 1.0}),
    ]
    log = tmp_path / "exp.jsonl"
    with open(log, "w", encoding="utf-8") as fh:
        for r in records:
            fh.write(json.dumps(r) + "\n")
    return str(log)


def _build(tmp_path) -> openpyxl.Workbook:
    from scripts.reporting.export_full_results_deck import (  # type: ignore
        build_deck, DEFAULT_DECK,
    )

    out = tmp_path / DEFAULT_DECK
    build_deck(_log(tmp_path), "config/taxonomy.yaml", str(out))
    return openpyxl.load_workbook(out)


def test_deck_has_19_slides_with_banner_and_footer(tmp_path):
    wb = _build(tmp_path)
    assert len(wb.sheetnames) == 19
    assert wb.sheetnames[0] == "01. COVER"
    for name in wb.sheetnames[1:]:
        ws = wb[name]
        assert "slide" in str(ws.cell(row=36, column=1).value)  # footer
        assert ws.page_setup.orientation == "landscape"


def test_extraction_lineage_rows_and_champion(tmp_path):
    wb = _build(tmp_path)
    ws = wb["03. EXTRACTION · Run lineage — "]
    rows = [r for r in ws.iter_rows(min_row=4, values_only=True) if r[1]]
    names = [str(r[1]) for r in rows]
    assert any("contracts_specialist_v2" in n for n in names)
    assert any("contracts_specialist_v32" in n for n in names)
    v32 = next(r for r in rows if "contracts_specialist_v32" in str(r[1]))
    assert v32[3] == 509
    assert abs(float(v32[4]) - 0.8807) < 1e-4


def test_sweep_slide_shows_all_models(tmp_path):
    wb = _build(tmp_path)
    ws = wb["11. SORTER SWEEP · Full model s"]
    rows = [r for r in ws.iter_rows(min_row=4, values_only=True) if r[1]]
    assert len(rows) == 2  # both synthetic sorter_v13 runs
    assert any("deepseek" in str(r[2]).lower() for r in rows)  # model column
    ds = next(r for r in rows if "deepseek" in str(r[2]).lower())
    assert abs(float(ds[4]) - 0.9528) < 1e-4  # subtype accuracy column


def test_legalbench_log_slides_and_summary(tmp_path):
    wb = _build(tmp_path)
    ws = wb["15. LEGALBENCH · Performance lo"]
    rows = [r for r in ws.iter_rows(min_row=4, values_only=True) if r[1]]
    v2 = next(r for r in rows if "task_v2_test" in str(r[1]))
    assert v2[3] == 94
    assert abs(float(v2[4]) - 0.8830) < 1e-4  # exact match column
    ws16 = wb["16. LEGALBENCH · Performance lo"]
    rows16 = [r for r in ws16.iter_rows(min_row=4, values_only=True) if r[1]]
    assert any("audit_rights" in str(r[1]) for r in rows16)
    summary = wb["18. LEGALBENCH · Per-task resul"]
    text = " ".join(str(c) for row in summary.iter_rows(values_only=True)
                    for c in row if c is not None)
    assert "task_v4_audit_rights" in text or "audit_rights" in text
